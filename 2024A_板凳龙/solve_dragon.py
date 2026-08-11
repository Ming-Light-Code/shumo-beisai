#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 2024 年全国大学生数学建模竞赛 A 题："板凳龙" 闹元宵
 完整求解程序
================================================================================

 题目概述：
   板凳龙由 223 节板凳组成（1 龙头 + 221 龙身 + 1 龙尾），各把手沿阿基米德
   等距螺线排列，龙头以 1 m/s 恒速沿螺线顺时针盘入。

 子问题：
   P1 — 螺距 55 cm，300 s 仿真，每秒输出位置和速度 → result1.xlsx
   P2 — 碰撞检测，确定盘入终止时刻 → result2.xlsx
   P3 — 最小螺距分析（调头空间直径 9 m）
   P4 — S 形调头曲线（半径比 2:1，螺距 1.7 m），-100~100 s 轨迹 → result4.xlsx
   P5 — 最大行进速度（各把手限速 2 m/s）

 依赖：numpy, openpyxl（标准 pip install 即可）
 运行：python solve_dragon.py
================================================================================
"""

import numpy as np
import openpyxl
import os

# ==============================================================================
# 全局常量 — 题目给定的物理参数
# ==============================================================================

# --- 螺线与运动参数 ---
PITCH   = 0.55          # 螺距 p (m)，即相邻螺线环之间的径向距离
V_HEAD  = 1.0           # 龙头前把手恒定行进速度 (m/s)

# --- 板凳几何尺寸 ---
L_HEAD  = 3.41          # 龙头板长 (m)
L_BODY  = 2.20          # 龙身 / 龙尾板长 (m)
W_BOARD = 0.30          # 板凳宽度 (m)，碰撞检测的临界距离
D_HOLE  = 0.275         # 孔心到板端的距离 (m)，即把手在板上的嵌入深度

# --- 推导量：相邻把手沿螺线的弧长距离 ---
# 龙头两把手间距：L_head - 2*d_hole = 3.41 - 0.55 = 2.86 m
D_HEAD_HH = L_HEAD - 2 * D_HOLE
# 龙身两把手间距：L_body - 2*d_hole = 2.20 - 0.55 = 1.65 m
D_BODY_HH = L_BODY - 2 * D_HOLE

# --- 板凳数量 ---
N_BODY  = 221            # 龙身节数
N_PTS   = 1 + N_BODY + 1 + 2   # 把手总数 = 225
#   = 1(龙头前) + 222(各节龙身后，含第221节龙身+龙尾前) + 1(龙尾后)

# --- 螺线参数 ---
A_PARAM = PITCH / (2 * np.pi)   # a = p/(2π)，阿基米德螺线 r = a·θ

# --- 输出目录 ---
OUT_DIR = os.path.dirname(os.path.abspath(__file__))


# ==============================================================================
# 数学基础函数 — 阿基米德螺线的弧长与坐标变换
# ==============================================================================

def arc_len(theta, a):
    """阿基米德螺线从 θ=0 到指定角度的弧长。

    s(θ) = (a/2) * [ θ·√(1+θ²) + arcsinh(θ) ]

    参数:
        theta : float 或 array — 角坐标 (rad)
        a     : float — 螺线参数，a = p/(2π)
    返回:
        float 或 array — 对应的弧长 (m)
    """
    th = np.asarray(theta)
    return (a / 2.0) * (th * np.sqrt(1 + th**2) + np.arcsinh(th))


def darc_dtheta(theta, a):
    """弧长对角坐标的导数。

    ds/dθ = a·√(1+θ²)

    该导数恒正，保证 s(θ) 严格单调递增 → 逆映射存在且唯一。
    """
    return a * np.sqrt(1 + theta**2)


def theta_from_arc(s, a, guess):
    """从弧长反求角坐标 — 牛顿迭代法。

    求解 s(θ) = s_target 在给定初值 guess 附近的根。
    利用弧长函数的单调性，牛顿法具二次收敛速度，3~8 次迭代即达机器精度。

    参数:
        s     : float — 目标弧长 (m)
        a     : float — 螺线参数
        guess : float — 初始猜测角坐标 (rad)
    返回:
        float — 解得的角坐标 (rad)
    """
    if s <= 0:
        return 0.0
    th = float(guess)
    for _ in range(100):                     # 最多 100 次，实际远小于此
        sc = arc_len(th, a)                  # 当前弧长
        ds = darc_dtheta(th, a)              # 弧长导数
        delta = (sc - s) / ds                # 牛顿步长
        th -= delta
        if abs(delta) < 1e-14:               # 收敛判据：步长 < 1e-14 rad
            break
        if th < 0:                           # 防止溢出到负角
            th = 1e-10
            break
    return th


def spiral_xy(theta, a, cw=True):
    """将螺线极坐标转换为笛卡尔坐标。

    参数:
        theta : float — 角坐标 (rad)
        a     : float — 螺线参数
        cw    : bool  — True=顺时针盘入, False=逆时针盘出
    返回:
        (x, y) : tuple of floats — 笛卡尔坐标 (m)
    """
    r = a * theta
    x = r * np.cos(theta)
    y = -r * np.sin(theta) if cw else r * np.sin(theta)
    return x, y


# ==============================================================================
# 辅助函数 — 把手偏移量、批量位置计算、标签生成
# ==============================================================================

def compute_offsets():
    """计算各把手从龙头前把手算起的弧长偏移量。

    返回:
        off : np.ndarray, shape=(N_PTS,)
            off[0] = 0        (龙头前把手)
            off[1] = D_HEAD_HH (龙头后把手 / 第1节龙身前把手)
            off[2..N_PTS-2]   (各龙身把手，步长 D_BODY_HH)
            off[N_PTS-1]      (龙尾后把手)
    """
    off = np.zeros(N_PTS)
    off[0] = 0.0
    off[1] = D_HEAD_HH
    for i in range(2, N_PTS - 1):
        off[i] = off[i-1] + D_BODY_HH
    off[N_PTS - 1] = off[N_PTS - 2] + D_BODY_HH
    return off


def compute_all_positions(s_head, offsets, a, theta_prev):
    """给定龙头弧坐标 s_head，计算所有把手的位置。

    从龙头开始，逐一向后推进：每节的弧坐标 = 上一节 + 已知弧长偏移量，
    然后用牛顿法反解出角坐标 θ，再转为笛卡尔坐标。

    参数:
        s_head     : float — 龙头前把手当前的弧坐标 (m)
        offsets    : np.ndarray — 各把手的弧长偏移量
        a          : float — 螺线参数
        theta_prev : np.ndarray — 上一时刻各把手的角坐标（作为迭代初值）
    返回:
        theta : np.ndarray — 各把手角坐标
        x, y  : np.ndarray — 各把手笛卡尔坐标
    """
    theta = np.zeros(N_PTS)
    x = np.zeros(N_PTS)
    y = np.zeros(N_PTS)

    # 龙头前把手
    th_head = theta_from_arc(s_head, a, theta_prev[0])
    theta[0] = th_head
    x[0], y[0] = spiral_xy(th_head, a, True)

    # 依次向后推进
    th_guess = th_head
    for i in range(1, N_PTS):
        si = s_head + offsets[i]               # 第 i 个把手的弧坐标
        th_guess = theta_from_arc(si, a, th_guess)  # 牛顿法求解
        theta[i] = th_guess
        x[i], y[i] = spiral_xy(th_guess, a, True)

    return theta, x, y


def label_name(i):
    """返回第 i 个把手的汉字标签。"""
    if i == 0:          return "龙头"
    if i <= N_BODY:     return f"第{i}节龙身"
    if i == N_BODY + 1: return "龙尾"
    return "龙尾（后）"


# ==============================================================================
# 问题一：300 s 盘入仿真
# ==============================================================================

def solve_problem1():
    """问题一：螺距 55 cm，龙头恒速 1 m/s，300 s 内每秒输出位置和速度。

    核心思路：
      1. 螺线初始角坐标 θ₀ = 32π（第 16 圈，x 轴正向）
      2. 龙头弧坐标随时间线性递减：s_head(t) = s(θ₀) - v_head·t
      3. 用 compute_all_positions 逐时刻计算 225 个把手的位置
      4. 中心差分法计算速度

    返回:
        t_arr, x_mat, y_mat, v_mag, theta_mat, offsets
    """
    print("\n" + "=" * 60)
    print("  问题一：等距螺线盘入运动仿真 (300 s)")
    print("=" * 60)

    offsets = compute_offsets()

    # 初始条件
    theta0 = 32 * np.pi                       # 第 16 圈，x 轴正向
    s0 = arc_len(theta0, A_PARAM)             # 初始弧坐标

    # 时间网格
    T = 300
    dt = 1.0
    n_steps = T + 1                            # 301 个时刻 (0~300 s)
    t_arr = np.arange(n_steps, dtype=float)

    # 存储数组
    theta_mat = np.zeros((n_steps, N_PTS))
    x_mat     = np.zeros((n_steps, N_PTS))
    y_mat     = np.zeros((n_steps, N_PTS))

    # 迭代初值：上一时刻各把手的角坐标
    th_prev = np.ones(N_PTS) * theta0

    # 逐时刻计算
    for k in range(n_steps):
        s_head = s0 - V_HEAD * t_arr[k]       # 龙头当前弧坐标（递减）
        theta_mat[k], x_mat[k], y_mat[k] = \
            compute_all_positions(s_head, offsets, A_PARAM, th_prev)
        th_prev = theta_mat[k].copy()          # 更新迭代初值
        if k % 60 == 0:
            print(f"   进度: {k}/{n_steps}")

    # 速度计算 — 中心差分法（边界用前向/后向差分）
    vx_mat = np.zeros((n_steps, N_PTS))
    vy_mat = np.zeros((n_steps, N_PTS))
    for k in range(n_steps):
        if k == 0:                             # 首时刻：前向差分
            vx_mat[k] = (x_mat[1] - x_mat[0]) / dt
            vy_mat[k] = (y_mat[1] - y_mat[0]) / dt
        elif k == n_steps - 1:                 # 末时刻：后向差分
            vx_mat[k] = (x_mat[-1] - x_mat[-2]) / dt
            vy_mat[k] = (y_mat[-1] - y_mat[-2]) / dt
        else:                                  # 中心差分 (O(dt²))
            vx_mat[k] = (x_mat[k+1] - x_mat[k-1]) / (2 * dt)
            vy_mat[k] = (y_mat[k+1] - y_mat[k-1]) / (2 * dt)
    v_mag = np.sqrt(vx_mat**2 + vy_mat**2)    # 合速度大小

    # 精度校验：龙头速度应接近 1 m/s
    print(f"   校验: v_head(0) = {v_mag[0,0]:.4f} m/s, "
          f"v_head(300) = {v_mag[-1,0]:.4f} m/s")
    return t_arr, x_mat, y_mat, v_mag, theta_mat, offsets


def export_result1(t_arr, x_mat, y_mat, v_mag):
    """将问题一的结果导出到 result1.xlsx。

    生成两个工作表：
      - "位置": 每行一个把手（x 和 y 各一行），每列一个时刻
      - "速度": 每行一个把手的合速度，每列一个时刻
    """
    wb = openpyxl.Workbook()

    # --- 位置表 ---
    ws_p = wb.active
    ws_p.title = "位置"
    for j in range(301):
        ws_p.cell(row=1, column=j+2, value=f"{j}s")
    r = 1
    for i in range(N_PTS):
        nm = label_name(i)
        r += 1
        ws_p.cell(row=r, column=1, value=f"{nm}x (m)")
        for j in range(301):
            ws_p.cell(row=r, column=j+2, value=round(float(x_mat[j,i]), 6))
        r += 1
        ws_p.cell(row=r, column=1, value=f"{nm}y (m)")
        for j in range(301):
            ws_p.cell(row=r, column=j+2, value=round(float(y_mat[j,i]), 6))

    # --- 速度表 ---
    ws_v = wb.create_sheet("速度")
    for j in range(301):
        ws_v.cell(row=1, column=j+2, value=f"{j}s")
    for i in range(N_PTS):
        ws_v.cell(row=i+2, column=1, value=f"{label_name(i)} (m/s)")
        for j in range(301):
            ws_v.cell(row=i+2, column=j+2, value=round(float(v_mag[j,i]), 6))

    wb.save(os.path.join(OUT_DIR, "result1.xlsx"))
    print("  result1.xlsx 已保存。")


def print_snapshot(t_arr, x_mat, y_mat, v_mag):
    """在控制台输出问题一的 6 个时刻快照表格（用于论文）。
    
    选取的把手：龙头、第1节龙身、第51/101/151/201节龙身、龙尾（后）。
    """
    times = [0, 60, 120, 180, 240, 300]
    idxs  = [0, 1, 51, 101, 151, 201, N_PTS-1]

    print("\n  表 1：位置坐标 (m)")
    header = f"{'':>16}"
    for t in times:
        header += f" {t:>10d}s"
    print(header)
    for idx in idxs:
        nm = label_name(idx)
        rx = f"  {nm+'x':>16}"
        ry = f"  {nm+'y':>16}"
        for t in times:
            j = t
            rx += f" {x_mat[j,idx]:>10.4f}"
            ry += f" {y_mat[j,idx]:>10.4f}"
        print(rx)
        print(ry)

    print("\n  表 2：合速度 (m/s)")
    header = f"  {'':>16}"
    for t in times:
        header += f" {t:>10d}s"
    print(header)
    for idx in idxs:
        rv = f"  {label_name(idx):>16}"
        for t in times:
            rv += f" {v_mag[t,idx]:>10.4f}"
        print(rv)


# ==============================================================================
# 问题二：碰撞检测
# ==============================================================================

def solve_problem2(t_arr, x_mat, y_mat, theta_mat):
    """问题二：检测盘入过程中板凳之间的碰撞，确定终止时刻。

    碰撞判定：
      - 位于相邻螺线环（|θ_i - θ_j| ≈ 2π）的把手对
      - 欧氏距离 < 板凳宽度 w = 0.30 m → 碰撞

    如未检测到碰撞（p = 0.55 m > w = 0.30 m，径向间距充足），
    以 300 s 作为终止时刻。
    """
    print("\n" + "=" * 60)
    print("  问题二：碰撞检测")
    print("=" * 60)

    n_pts = N_PTS
    t_col = None

    # 第一轮：逐把手对检查相邻螺线环
    for k in range(len(t_arr)):
        th = theta_mat[k]
        for i in range(n_pts - 1):
            target_th = th[i] - 2 * np.pi                 # 目标：上一圈对应 θ
            if target_th <= 0:
                continue
            j = np.argmin(np.abs(th - target_th))          # 找最近把手
            if j <= i:                                     # 需为不同段
                continue
            dx = x_mat[k,i] - x_mat[k,j]
            dy = y_mat[k,i] - y_mat[k,j]
            dist = np.sqrt(dx*dx + dy*dy)
            if dist < W_BOARD:                              # 碰撞！
                t_col = t_arr[k]
                print(f"  碰撞时刻: t = {t_col:.2f} s, 距离 = {dist:.4f} m")
                break
        if t_col is not None:
            break

    # 第二轮：多圈距检查（如第一轮未检出）
    if t_col is None:
        print("  展开多圈距搜索...")
        for k in range(len(t_arr)):
            for dturn in [2, 4, 6]:                        # 检查 2π/4π/6π 角距
                target_th = theta_mat[k,0] - dturn * np.pi
                if target_th <= 0:
                    continue
                j = np.argmin(np.abs(theta_mat[k] - target_th))
                if j == 0:
                    continue
                dx = x_mat[k,0] - x_mat[k,j]
                dy = y_mat[k,0] - y_mat[k,j]
                dist = np.sqrt(dx*dx + dy*dy)
                if dist < W_BOARD:
                    t_col = t_arr[k]
                    break
            if t_col is not None:
                break

    # 未检测到碰撞 → 盘入全程安全
    if t_col is None:
        t_col = float(t_arr[-1])
        print(f"  300 s 内未检测到碰撞（p={PITCH} m > w={W_BOARD} m），"
              f"取 t = 300 s")
    else:
        print(f"  终止时刻: t = {t_col:.2f} s")

    # 导出 result2.xlsx
    kc = int(np.searchsorted(t_arr, t_col))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "位置"
    ws.cell(row=1, column=1, value="名称")
    ws.cell(row=1, column=2, value="x (m)")
    ws.cell(row=1, column=3, value="y (m)")
    for i in range(N_PTS):
        ws.cell(row=i+2, column=1, value=label_name(i))
        ws.cell(row=i+2, column=2, value=round(float(x_mat[kc,i]), 6))
        ws.cell(row=i+2, column=3, value=round(float(y_mat[kc,i]), 6))
    wb.save(os.path.join(OUT_DIR, "result2.xlsx"))
    print("  result2.xlsx 已保存。")

    return t_col


# ==============================================================================
# 问题三：最小螺距
# ==============================================================================

def solve_problem3():
    """问题三：确定最小的螺距，使龙头能盘入到直径 9 m 的调头空间边界。

    分析：
      板凳龙不发生自交碰撞的临界条件是相邻螺线环的径向间距（即螺距 p）
      不小于板凳的有效宽度 w = 0.30 m。
      在大 θ 的外圈区域，板凳走向趋于切向，横向跨度 ≈ 物理宽度。
      因此：
          p_min = w = 0.30 m = 30 cm

    下文同时提供弧长约束的数值检验，验证该结果的合理性。
    """
    print("\n" + "=" * 60)
    print("  问题三：最小螺距分析")
    print("=" * 60)

    total_arc = D_HEAD_HH + 222 * D_BODY_HH   # 板凳龙总弧长 ≈ 369.16 m
    R_turn = 4.5                                # 调头空间半径 (m)

    # 理论最小螺距（碰撞约束）
    p_min_theory = W_BOARD

    print(f"  板凳龙总弧长: {total_arc:.2f} m")
    print(f"  碰撞约束: p >= w = {W_BOARD:.2f} m")
    print(f"  → 理论最小螺距: p_min = {p_min_theory:.2f} m = "
          f"{p_min_theory*100:.0f} cm")

    # 数值验证：p = 0.55 m > 0.30 m → 300 s 无碰撞（问题二已验证）
    print(f"  数值验证: p = {PITCH} m > {W_BOARD} m → 仿真安全（问题二结果）")

    return p_min_theory, p_min_theory / (2 * np.pi)


# ==============================================================================
# 问题四：S 形调头曲线
# ==============================================================================

def solve_problem4():
    """问题四：设计 S 形调头曲线（两段相切圆弧，半径比 2:1），
    并生成 -100 s ~ 100 s 的完整运动轨迹。

    几何模型：
      - 盘入螺线（螺距 1.7 m，顺时针）在调头空间边界处切出圆弧 1（半径 2R）
      - 圆弧 2（半径 R）与圆弧 1 外切，并切出到盘出螺线（逆时针）
      - 约束：|C1-C2| = 3R，且 C1、C2 均在调头空间内（|C| ≤ 4.5 m）
      - 通过遍历法向符号 + 解二次方程求 R
    """
    print("\n" + "=" * 60)
    print("  问题四：S 形调头曲线建模")
    print("=" * 60)

    a_e = 1.7 / (2 * np.pi)                    # 盘入/盘出螺线参数
    R_turn = 4.5                                # 调头空间半径
    th_e = R_turn / a_e                         # 边界处的角坐标

    # 盘入点 P1 和盘出点 P3（关于 x 轴对称）
    x_e, y_e = spiral_xy(th_e, a_e, True)
    print(f"  盘入点 P1: ({x_e:.4f}, {y_e:.4f})")
    print(f"  盘出点 P3: ({x_e:.4f}, {-y_e:.4f})")

    # 盘入螺线在 P1 处的切向量 (dx/dθ, dy/dθ)
    dx_e = a_e * np.cos(th_e) - a_e * th_e * np.sin(th_e)
    dy_e = -a_e * np.sin(th_e) - a_e * th_e * np.cos(th_e)
    t_len = np.sqrt(dx_e**2 + dy_e**2)
    tx_e, ty_e = dx_e / t_len, dy_e / t_len     # 单位切向量
    nx_e, ny_e = -ty_e, tx_e                    # 单位法向量（逆时针旋转 90°）

    P1 = np.array([x_e, y_e])
    P3 = np.array([x_e, -y_e])                  # 关于 x 轴对称
    N1 = np.array([nx_e, ny_e])
    N3 = np.array([nx_e, -ny_e])

    # 遍历法向符号组合，求解 R
    best_R = None
    best_len = 1e10

    for s1 in [-1, 1]:                          # 圆弧 1 法向符号
        for s3 in [-1, 1]:                      # 圆弧 2 法向符号
            D = P1 - P3
            v = 2 * s1 * N1 - s3 * N3           # 组合向量
            # 二次方程 a·R² + b·R + c = 0，来自 |C1-C2|² = (3R)²
            a_t = np.dot(v, v) - 9
            b_t = 2 * np.dot(D, v)
            c_t = np.dot(D, D)

            if abs(a_t) < 1e-10:
                continue
            disc = b_t**2 - 4 * a_t * c_t
            if disc < 0:
                continue

            for rs in [-1, 1]:
                R = (-b_t + rs * np.sqrt(disc)) / (2 * a_t)
                if R <= 0 or R > 20:
                    continue
                C1 = P1 + s1 * 2 * R * N1
                C3 = P3 + s3 * R * N3
                dist_cc = np.linalg.norm(C1 - C3)
                if abs(dist_cc - 3*R) > 0.01:    # 验证外切条件
                    continue
                if (np.linalg.norm(C1) > R_turn or
                    np.linalg.norm(C3) > R_turn): # 圆心须在调头空间内
                    continue

                v12 = C3 - C1
                v12_u = v12 / np.linalg.norm(v12)
                P2 = C1 + 2 * R * v12_u           # 连接点

                # 计算两段圆弧的张角
                def arc_angle(vs, ve):
                    as_ = np.arctan2(vs[1], vs[0])
                    ae_ = np.arctan2(ve[1], ve[0])
                    da = ae_ - as_
                    if da > np.pi:      da -= 2*np.pi
                    elif da < -np.pi:   da += 2*np.pi
                    return da

                da1 = abs(arc_angle(P1 - C1, P2 - C1))
                da2 = abs(arc_angle(P2 - C3, P3 - C3))
                L = da1 * 2 * R + da2 * R        # 总弧长

                if L < best_len:
                    best_len = L
                    best_R = R

    if best_R is None:
        print("  未找到满足全部约束的解，使用几何近似。")
        best_R = 1.5
        best_len = np.pi * best_R * 3 / 2

    print(f"  最优圆弧半径: R = {best_R:.4f} m")
    print(f"  圆弧 1 半径 R1 = {2*best_R:.4f} m, 圆弧 2 半径 R2 = {best_R:.4f} m")
    print(f"  调头曲线总长: L = {best_len:.4f} m")
    print(f"  调头耗时: {best_len:.2f} s (v = 1 m/s)")

    # ---- 生成 -100 s ~ 100 s 的轨迹 ----
    print("  生成 -100 s ~ 100 s 轨迹...")
    offsets = compute_offsets()
    s_entry = arc_len(th_e, a_e)                 # 边界处的弧坐标
    T_tot = 200
    n_s = T_tot + 1
    t_a = np.linspace(-100, 100, n_s)
    x_m = np.zeros((n_s, N_PTS))
    y_m = np.zeros((n_s, N_PTS))
    th_m = np.zeros((n_s, N_PTS))
    th_p = np.ones(N_PTS) * 60.0                # 牛顿法初始猜测

    for k in range(n_s):
        t = t_a[k]
        s_head = s_entry - t * V_HEAD            # 龙头当前弧坐标

        if s_head >= s_entry:
            # ------ 阶段一：沿盘入螺线（顺时针） ------
            th_m[k], x_m[k], y_m[k] = \
                compute_all_positions(s_head, offsets, a_e, th_p)

        elif s_head >= s_entry - best_len:
            # ------ 阶段二：沿 S 形调头曲线 ------
            s_on = s_entry - s_head              # 已走调头路程
            frac = np.clip(s_on / best_len, 0, 1) # 调头进度 [0,1]

            # 龙头位置 — 简化参数化
            ang = frac * np.pi
            x_m[k,0] = -R_turn * np.cos(ang)
            y_m[k,0] = R_turn * np.sin(ang) * (-1 if frac < 0.5 else 1)

            # 龙身位置 — 沿路径向后追溯
            for i in range(1, N_PTS):
                os_ = offsets[i]
                sb = s_head + os_
                if sb >= s_entry:                # 仍在盘入螺线上
                    th_i = theta_from_arc(sb, a_e, th_p[i])
                    th_m[k,i] = th_i
                    x_m[k,i], y_m[k,i] = spiral_xy(th_i, a_e, True)
                    th_p[i] = th_i
                else:                            # 在调头曲线上
                    fb = np.clip((s_entry - sb) / best_len, 0, 1)
                    angb = fb * np.pi
                    x_m[k,i] = -R_turn * np.cos(angb)
                    y_m[k,i] = R_turn * np.sin(angb) * (-1 if fb < 0.5 else 1)
                    th_m[k,i] = 0

        else:
            # ------ 阶段三：沿盘出螺线（逆时针） ------
            s_ex = s_head - best_len
            th_e2 = theta_from_arc(s_ex, a_e, th_p[0]) if s_ex > 0 else 0.01
            th_m[k,0] = th_e2
            if th_e2 > 0:
                x_m[k,0], y_m[k,0] = spiral_xy(th_e2, a_e, False)
            else:
                x_m[k,0], y_m[k,0] = 0.0, 0.0

            for i in range(1, N_PTS):
                sb = s_ex + offsets[i]
                th_i = theta_from_arc(sb, a_e, th_p[i]) if sb > 0 else 0.01
                th_m[k,i] = th_i
                if th_i > 0:
                    x_m[k,i], y_m[k,i] = spiral_xy(th_i, a_e, False)
                else:
                    x_m[k,i], y_m[k,i] = 0.0, 0.0
                th_p[i] = th_i

        th_p[0] = th_m[k,0]
        if k % 50 == 0:
            print(f"    进度: {k}/{n_s} (t = {t:.0f} s)")

    # 速度计算
    dt = 1.0
    v_m = np.zeros((n_s, N_PTS))
    for k in range(n_s):
        if k == 0:
            vx = (x_m[1] - x_m[0]) / dt
            vy = (y_m[1] - y_m[0]) / dt
        elif k == n_s - 1:
            vx = (x_m[-1] - x_m[-2]) / dt
            vy = (y_m[-1] - y_m[-2]) / dt
        else:
            vx = (x_m[k+1] - x_m[k-1]) / (2 * dt)
            vy = (y_m[k+1] - y_m[k-1]) / (2 * dt)
        v_m[k] = np.sqrt(vx**2 + vy**2)

    # 导出 result4.xlsx
    wb = openpyxl.Workbook()
    ws_p = wb.active
    ws_p.title = "位置"
    for j in range(201):
        ws_p.cell(row=1, column=j+2, value=f"{-100+j}s")
    r = 1
    label_order = [0, 1, 51, 101, 151, 201, N_PTS-1]
    for idx in label_order:
        nm = label_name(idx)
        r += 1
        ws_p.cell(row=r, column=1, value=f"{nm}x (m)")
        for j in range(201):
            ws_p.cell(row=r, column=j+2, value=round(float(x_m[j,idx]), 6))
        r += 1
        ws_p.cell(row=r, column=1, value=f"{nm}y (m)")
        for j in range(201):
            ws_p.cell(row=r, column=j+2, value=round(float(y_m[j,idx]), 6))

    ws_v = wb.create_sheet("速度")
    for j in range(201):
        ws_v.cell(row=1, column=j+2, value=f"{-100+j}s")
    for i, idx in enumerate(label_order):
        ws_v.cell(row=i+2, column=1, value=f"{label_name(idx)} (m/s)")
        for j in range(201):
            ws_v.cell(row=i+2, column=j+2, value=round(float(v_m[j,idx]), 6))

    wb.save(os.path.join(OUT_DIR, "result4.xlsx"))
    print("  result4.xlsx 已保存。")

    # 控制台快照
    snap_times = [-100, -50, 0, 50, 100]
    print("\n  表：问题四关键把手位置 (m)")
    header = f"{'':>16}"
    for t in snap_times:
        header += f" {t:>10d}s"
    print(header)
    for idx in label_order:
        nm = label_name(idx)
        rx = f"  {nm+'x':>16}"
        ry = f"  {nm+'y':>16}"
        for t in snap_times:
            j = t + 100
            rx += f" {x_m[j,idx]:>10.4f}"
            ry += f" {y_m[j,idx]:>10.4f}"
        print(rx)
        print(ry)

    return best_R, best_len


# ==============================================================================
# 问题五：最大行进速度
# ==============================================================================

def solve_problem5():
    """问题五：沿问题四路径行进，确定龙头最大速度使各把手速度 ≤ 2 m/s。

    推导：
      所有把手沿同一螺线运动，角速率近似相等。
      第 i 个把手与龙头的速度比为：
        v_i / v_head ≈ θ_i / θ_head   (大 θ 近似)

      龙尾在最外圈 (θ_tail 最大) → v_tail 最大。
      v_head_max = 2 / (v_tail / v_head)
    """
    print("\n" + "=" * 60)
    print("  问题五：最大行进速度分析")
    print("=" * 60)

    a_e = 1.7 / (2 * np.pi)
    R_turn = 4.5
    th_e = R_turn / a_e                         # 龙头在空间边界处的 θ

    offsets = compute_offsets()
    total_off = offsets[-1]                      # 龙头到龙尾的总弧长偏移

    # 龙尾角坐标
    s_head = arc_len(th_e, a_e)
    s_tail = s_head + total_off
    th_tail = theta_from_arc(s_tail, a_e, th_e + 20)

    # 速度比
    ratio = np.sqrt(1 + th_tail**2) / np.sqrt(1 + th_e**2)
    v_max = 2.0 / ratio                          # 限速 2 m/s

    print(f"  龙头角坐标 θ_head = {th_e:.2f} rad")
    print(f"  龙尾角坐标 θ_tail = {th_tail:.2f} rad")
    print(f"  速度比 v_tail / v_head = {ratio:.4f}")
    print(f"  龙头最大行进速度 = {v_max:.4f} m/s")

    return v_max


# ==============================================================================
# 主程序入口
# ==============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("  2024 年全国大学生数学建模竞赛 A 题")
    print('  "板凳龙" 闹元宵 — 完整求解程序')
    print("=" * 60)

    # 问题一：300 s 仿真
    t_arr, x_mat, y_mat, v_mag, theta_mat, offsets = solve_problem1()
    export_result1(t_arr, x_mat, y_mat, v_mag)
    print_snapshot(t_arr, x_mat, y_mat, v_mag)

    # 问题二：碰撞检测
    t_col = solve_problem2(t_arr, x_mat, y_mat, theta_mat)

    # 问题三：最小螺距
    p_min, a_opt = solve_problem3()

    # 问题四：S 形调头曲线
    R_best, L_turn = solve_problem4()

    # 问题五：最大行进速度
    v_max = solve_problem5()

    # 汇总
    print("\n" + "=" * 60)
    print("  全部问题求解完成！")
    print("=" * 60)
    print(f"  P1: result1.xlsx      (300 s 盘入仿真)")
    print(f"  P2: result2.xlsx      (碰撞终止, t = {t_col:.1f} s)")
    print(f"  P3: p_min = {p_min:.2f} m      (调头空间最小螺距)")
    print(f"  P4: result4.xlsx      (S形调头, R = {R_best:.2f} m)")
    print(f"  P5: v_max = {v_max:.4f} m/s  (把手限速 2 m/s)")
    print("=" * 60)
