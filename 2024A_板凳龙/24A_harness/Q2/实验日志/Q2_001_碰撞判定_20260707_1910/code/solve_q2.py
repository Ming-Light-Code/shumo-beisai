"""Q2 碰撞判定：复用问1螺线位置模型，OBB-SAT全穷举判碰，二分求终止时刻。

依据 Q2/实现方案/Q2_v1_碰撞判定方案.md。
输出：result2.xlsx（官方模板单sheet，终止时刻状态）、论文表格、粗扫描clearance曲线数据。
"""
import os
import numpy as np
from scipy.optimize import brentq

# ----------------------------- 常量（复用问1） -----------------------------
P = 0.55
B = P / (2.0 * np.pi)
V0 = 1.0
THETA0_INIT = 32.0 * np.pi
N_NODE = 224
L_HEAD = 3.41 - 2 * 0.275
L_BODY = 2.20 - 2 * 0.275
D = np.empty(N_NODE)
D[0] = np.nan
D[1] = L_HEAD
D[2:] = L_BODY
XTOL = 1e-12

# ----------------------------- Q2 新增常量：矩形几何 -----------------------------
N_BENCH = N_NODE - 1  # 223
HL = np.empty(N_BENCH)          # 半长 (m)
HL[0] = 3.41 / 2.0              # 龙头板 1.705 m
HL[1:] = 2.20 / 2.0             # 龙身/龙尾板 1.10 m
HW = 0.30 / 2.0                 # 半宽 0.15 m（全部一致）

# 候选对索引：j-i>=2（自动排除相邻共享铰接点对与自身）
_IU, _JU = np.triu_indices(N_BENCH, k=2)

PAPER_NODES = [0, 1, 51, 101, 151, 201, 223]
PAPER_NODE_LABELS = ["龙头", "第1节龙身", "第51节龙身", "第101节龙身",
                     "第151节龙身", "第201节龙身", "龙尾(后)"]


# --------------------------- 几何函数（同问1） ---------------------------
def xy(theta):
    r = B * theta
    return r * np.cos(theta), r * np.sin(theta)


def dxy(theta):
    return (B * (np.cos(theta) - theta * np.sin(theta)),
            B * (np.sin(theta) + theta * np.cos(theta)))


def arc_len(theta):
    s = np.sqrt(theta * theta + 1.0)
    return 0.5 * B * (theta * s + np.log(theta + s))


def solve_head_theta(t):
    target = arc_len(THETA0_INIT) - V0 * t
    g = lambda th: arc_len(th) - target
    return brentq(g, 0.0, THETA0_INIT, xtol=XTOL)


def solve_next_theta(theta_prev, d):
    xp, yp = xy(theta_prev)

    def F(th):
        x, y = xy(th)
        return (x - xp) ** 2 + (y - yp) ** 2 - d * d

    lo = theta_prev + 1e-9
    step = d / (B * np.sqrt(theta_prev ** 2 + 1.0))
    hi = theta_prev + step
    for _ in range(60):
        if F(hi) > 0:
            break
        hi = theta_prev + (hi - theta_prev) * 1.6
    else:
        raise RuntimeError(f"括号扩张失败 theta_prev={theta_prev}")
    return brentq(F, lo, hi, xtol=XTOL)


def chain_theta_from_theta0(theta0):
    """给定龙头极角，递推整条链，返回 theta[224]。（重构：与 t 解耦，供二分复用）"""
    theta = np.empty(N_NODE)
    theta[0] = theta0
    for i in range(1, N_NODE):
        theta[i] = solve_next_theta(theta[i - 1], D[i])
    return theta


def state_from_theta0(theta0):
    """给定龙头极角，返回 theta,x,y,vx,vy,v （速度只依赖 theta0，不显式依赖 t）。"""
    theta = chain_theta_from_theta0(theta0)
    x, y = xy(theta)
    dx, dy = dxy(theta)

    vx = np.empty(N_NODE)
    vy = np.empty(N_NODE)
    thd0 = -V0 / (B * np.sqrt(theta0 ** 2 + 1.0))
    vx[0] = dx[0] * thd0
    vy[0] = dy[0] * thd0
    for i in range(1, N_NODE):
        ddx = x[i] - x[i - 1]
        ddy = y[i] - y[i - 1]
        denom = ddx * dx[i] + ddy * dy[i]
        thdi = (ddx * vx[i - 1] + ddy * vy[i - 1]) / denom
        vx[i] = dx[i] * thdi
        vy[i] = dy[i] * thdi
    v = np.sqrt(vx ** 2 + vy ** 2)
    return theta, x, y, vx, vy, v


# ------------------------- Q2 核心：矩形 + SAT clearance -------------------------
def benches_from_xy(x, y):
    """由 224 节点坐标构造 223 节板凳的 (中心C, 杆向U, 法向N, 半长HL)。"""
    P_ = np.stack([x, y], axis=1)          # (224,2)
    p0 = P_[:-1]                            # (223,2)
    p1 = P_[1:]
    seg = p1 - p0
    length = np.linalg.norm(seg, axis=1)
    U = seg / length[:, None]
    N = np.stack([-U[:, 1], U[:, 0]], axis=1)
    C = 0.5 * (p0 + p1)
    return C, U, N


def clearance(theta0):
    """全局间隙 G(theta0)：>0 表示全部候选对分离(不碰)；<=0 表示至少一对碰撞。
    同时返回取得最小间隙的候选对下标 (i, j)（1-based 板凳编号）。
    """
    theta = chain_theta_from_theta0(theta0)
    x, y = xy(theta)
    C, U, N = benches_from_xy(x, y)

    Ci, Cj = C[_IU], C[_JU]
    Ui, Uj = U[_IU], U[_JU]
    Ni, Nj = N[_IU], N[_JU]
    HLi, HLj = HL[_IU], HL[_JU]
    cdiff = Cj - Ci

    def margin_for_axis(a):
        proj = np.abs(np.sum(cdiff * a, axis=1))
        ri = HLi * np.abs(np.sum(Ui * a, axis=1)) + HW * np.abs(np.sum(Ni * a, axis=1))
        rj = HLj * np.abs(np.sum(Uj * a, axis=1)) + HW * np.abs(np.sum(Nj * a, axis=1))
        return proj - (ri + rj)

    m1 = margin_for_axis(Ui)
    m2 = margin_for_axis(Ni)
    m3 = margin_for_axis(Uj)
    m4 = margin_for_axis(Nj)
    pair_margin = np.maximum(np.maximum(m1, m2), np.maximum(m3, m4))
    k = int(np.argmin(pair_margin))
    return float(pair_margin[k]), (int(_IU[k]) + 1, int(_JU[k]) + 1)  # 板凳编号 1-based


def bisect_theta0(theta0_hi, theta0_lo, xtol=1e-9, maxiter=100):
    """theta0_hi: G>0（未碰，t较小） theta0_lo: G<=0（已碰，t较大）。theta0 越小越靠内圈/越晚。
    返回收敛后的 theta0_hi（最后一个仍分离的临界值，即 Q2-A05 的临界边界）。
    """
    g_hi, _ = clearance(theta0_hi)
    g_lo, _ = clearance(theta0_lo)
    assert g_hi > 0 and g_lo <= 0, f"括号不满足符号相反: g_hi={g_hi}, g_lo={g_lo}"
    for _ in range(maxiter):
        if abs(theta0_hi - theta0_lo) < xtol:
            break
        mid = 0.5 * (theta0_hi + theta0_lo)
        g_mid, _ = clearance(mid)
        if g_mid > 0:
            theta0_hi = mid
        else:
            theta0_lo = mid
    return theta0_hi


# ------------------------------ 主流程 ------------------------------
def main():
    here = os.path.dirname(os.path.abspath(__file__))
    outdir = os.path.join(os.path.dirname(here), "outputs")
    os.makedirs(outdir, exist_ok=True)

    # Step 1: 粗扫描（步长1s，0..430s；龙头极角在 t≈442.59s 时耗尽到0，取430s留安全边际）
    # 记录 clearance 曲线，找到首次 G<=0 的区间。诊断确认首次碰撞发生在 t≈412~413s（晚于问1的300s窗口）。
    t_list = np.arange(0, 431)
    G = np.empty(len(t_list))
    crit_pairs = [None] * len(t_list)
    theta0_list = np.empty(len(t_list))
    first_collide_idx = None
    for j, t in enumerate(t_list):
        theta0 = solve_head_theta(t)
        theta0_list[j] = theta0
        g, pair = clearance(theta0)
        G[j] = g
        crit_pairs[j] = pair
        if g <= 0 and first_collide_idx is None:
            first_collide_idx = j

    if first_collide_idx is None:
        raise RuntimeError(f"{t_list[-1]}s 内粗扫描未检测到碰撞，需扩大时间范围复核（理论上界 t<{arc_len(THETA0_INIT):.3f}s）")

    j1 = first_collide_idx
    t_hi, t_lo = t_list[j1 - 1], t_list[j1]
    theta0_hi, theta0_lo = theta0_list[j1 - 1], theta0_list[j1]
    print(f"[coarse] 首次碰撞在 t∈[{t_hi},{t_lo}] s 之间，"
          f"G({t_hi})={G[j1-1]:.6f}, G({t_lo})={G[j1]:.6f}, 临界候选对(板凳编号)={crit_pairs[j1]}")

    # Step 2: 二分收敛（对 theta0）
    theta0_star = bisect_theta0(theta0_hi, theta0_lo, xtol=1e-9)
    t_star = arc_len(THETA0_INIT) - arc_len(theta0_star)
    g_star, pair_star = clearance(theta0_star)
    print(f"[bisect] theta0* = {theta0_star:.12f}, t* = {t_star:.6f} s, "
          f"G(theta0*) = {g_star:.3e}, 临界对(板凳编号) = {pair_star}")

    theta, x, y, vx, vy, v = state_from_theta0(theta0_star)

    # sanity: 终止时刻前 1 步应全不碰；此刻应临界(|G|很小)；再往前 1e-6 rad 应已碰
    g_before, _ = clearance(theta0_star + 1e-6)
    g_after, _ = clearance(theta0_star - 1e-6)
    print(f"[check] G(theta0*+1e-6, 更早)={g_before:.6e} (应>0) | "
          f"G(theta0*-1e-6, 更晚)={g_after:.6e} (应<=0) | |G(theta0*)|={abs(g_star):.3e}")

    _write_excel(outdir, x, y, v)
    _write_paper_row(outdir, t_star, theta0_star, x, y, v, pair_star)
    _save_coarse_curve(outdir, t_list, G, theta0_list)
    _save_snapshot(outdir, theta0_star, x, y)
    print("[done] outputs ->", outdir)


def _row_labels():
    labels = ["龙头"]
    labels += [f"第{k}节龙身" for k in range(1, 222)]
    labels += ["龙尾", "龙尾(后)"]
    return labels


def _find_template(name):
    d = os.path.dirname(os.path.abspath(__file__))
    for _ in range(8):
        cand = os.path.join(d, "题目", name)
        if os.path.exists(cand):
            return cand
        d = os.path.dirname(d)
    raise FileNotFoundError(f"未找到 题目/{name} 官方模板")


def _write_excel(outdir, x, y, v):
    import openpyxl
    tmpl = _find_template("result2.xlsx")
    wb = openpyxl.load_workbook(tmpl)
    ws = wb["Sheet1"]
    for i in range(N_NODE):
        ws.cell(2 + i, 2).value = round(float(x[i]), 6)
        ws.cell(2 + i, 3).value = round(float(y[i]), 6)
        ws.cell(2 + i, 4).value = round(float(v[i]), 6)
    path = os.path.join(outdir, "result2.xlsx")
    wb.save(path)
    print("[write]", path, "(filled official template)")


def _write_paper_row(outdir, t_star, theta0_star, x, y, v, pair_star):
    import csv
    with open(os.path.join(outdir, "paper_table_termination.csv"), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["终止时刻 t* (s)", f"{t_star:.6f}"])
        w.writerow(["龙头极角 theta0* (rad)", f"{theta0_star:.9f}"])
        w.writerow(["临界碰撞板凳对(1-based)", str(pair_star)])
        w.writerow([])
        w.writerow(["节点", "x (m)", "y (m)", "v (m/s)"])
        for node, lab in zip(PAPER_NODES, PAPER_NODE_LABELS):
            w.writerow([lab, f"{x[node]:.6f}", f"{y[node]:.6f}", f"{v[node]:.6f}"])
    print("[write] paper_table_termination.csv")


def _save_coarse_curve(outdir, t_list, G, theta0_list):
    np.savez(os.path.join(outdir, "coarse_scan.npz"), t=t_list, G=G, theta0=theta0_list)
    print("[write] coarse_scan.npz")


def _save_snapshot(outdir, theta0_star, x, y):
    np.savez(os.path.join(outdir, "termination_snapshot.npz"), theta0=theta0_star, x=x, y=y)
    print("[write] termination_snapshot.npz")


if __name__ == "__main__":
    main()
