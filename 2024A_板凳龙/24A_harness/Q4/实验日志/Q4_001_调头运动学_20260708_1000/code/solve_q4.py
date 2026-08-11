"""Q4 调头运动学主程序：龙头1m/s沿复合轨道，求-100..100s每秒224把手位置速度。

依据 Q4/实现方案/Q4_v1_调头运动学方案.md。复用 geometry_q4.Track。
"""
import os
import numpy as np
from scipy.optimize import brentq
import geometry_q4 as g

N_NODE = 224
L_HEAD = 3.41 - 2 * 0.275     # 2.86
L_BODY = 2.20 - 2 * 0.275     # 1.65
D = np.empty(N_NODE)
D[0] = np.nan
D[1] = L_HEAD
D[2:] = L_BODY

T_LIST = np.arange(-100, 101)          # -100..100 s
PAPER_TIMES = [-100, -50, 0, 50, 100]
PAPER_NODES = [0, 1, 51, 101, 151, 201, 223]
PAPER_NODE_LABELS = ["龙头", "第1节龙身", "第51节龙身", "第101节龙身",
                     "第151节龙身", "第201节龙身", "龙尾(后)"]


def solve_state(track, s_head):
    """给定龙头弧长坐标 s_head，返回 s[224], P[224,2], T[224,2], v[224]."""
    s = np.empty(N_NODE)
    P = np.empty((N_NODE, 2))
    T = np.empty((N_NODE, 2))
    s[0] = s_head
    P[0], T[0] = track.point_tangent(s_head)
    for i in range(1, N_NODE):
        s_prev = s[i - 1]
        Pprev = P[i - 1]
        Li = D[i]

        def gfun(si):
            return np.linalg.norm(Pprev - track.point(si)) - Li

        # 括号：si 在 [s_prev - span, s_prev - eps]，chord 随 (s_prev-si) 增而增（单调段）
        lo = s_prev - 2.2 * Li
        hi = s_prev - 1e-12
        # 保证变号；必要时向下扩张
        for _ in range(40):
            if gfun(lo) > 0:
                break
            lo = s_prev - (s_prev - lo) * 1.5
        si = brentq(gfun, lo, hi, xtol=1e-11, rtol=1e-14)
        s[i] = si
        P[i], T[i] = track.point_tangent(si)

    # 速度递推：sdot0=1
    sdot = np.empty(N_NODE)
    sdot[0] = 1.0
    for i in range(1, N_NODE):
        dP = P[i - 1] - P[i]
        num = np.dot(dP, T[i - 1])
        den = np.dot(dP, T[i])
        sdot[i] = num / den * sdot[i - 1]
    v = np.abs(sdot)                        # 弧长参数 => |T|=1 => 速度大小=|sdot|
    return s, P, T, v


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    outdir = os.path.join(os.path.dirname(here), "outputs")
    os.makedirs(outdir, exist_ok=True)

    track = g.Track(ratio=2.0, theta_max=75.0)
    print(f"[geom] LS={track.LS:.6f} m, R1={track.sc['R1']:.6f}, R2={track.sc['R2']:.6f}")

    nt = len(T_LIST)
    X = np.empty((N_NODE, nt))
    Y = np.empty((N_NODE, nt))
    V = np.empty((N_NODE, nt))
    max_resid = 0.0
    mono_ok = True
    for j, t in enumerate(T_LIST):
        s, P, T, v = solve_state(track, float(t))
        X[:, j] = P[:, 0]
        Y[:, j] = P[:, 1]
        V[:, j] = v
        for i in range(1, N_NODE):
            r = abs(np.linalg.norm(P[i - 1] - P[i]) - D[i])
            max_resid = max(max_resid, r)
        if not np.all(np.diff(s) < 0):
            mono_ok = False

    print(f"[check] 最大弦长残差 = {max_resid:.3e}")
    print(f"[check] s 沿链严格递减 = {mono_ok}")
    print(f"[check] 龙头速度 min/max = {V[0].min():.6f}/{V[0].max():.6f}")
    j0 = list(T_LIST).index(0)
    print(f"[check] t=0 龙头坐标 = ({X[0,j0]:.6f},{Y[0,j0]:.6f}) 应=P_in "
          f"({track.sc['P_in'][0]:.6f},{track.sc['P_in'][1]:.6f})")
    print(f"[check] 全场速度 min/max = {V.min():.6f}/{V.max():.6f}")

    _write_excel(outdir, X, Y, V)
    _write_paper_tables(outdir, X, Y, V)
    np.savez(os.path.join(outdir, "q4_fields.npz"), t=T_LIST, X=X, Y=Y, V=V)
    print("[done] outputs ->", outdir)


def _find_template(name):
    d = os.path.dirname(os.path.abspath(__file__))
    for _ in range(9):
        cand = os.path.join(d, "题目", name)
        if os.path.exists(cand):
            return cand
        d = os.path.dirname(d)
    raise FileNotFoundError(f"未找到 题目/{name}")


def _write_excel(outdir, X, Y, V):
    import openpyxl
    wb = openpyxl.load_workbook(_find_template("result4.xlsx"))
    ws1, ws2 = wb["位置"], wb["速度"]
    nt = X.shape[1]
    for i in range(N_NODE):
        for j in range(nt):
            ws1.cell(2 + 2 * i, 2 + j).value = round(float(X[i, j]), 6)
            ws1.cell(3 + 2 * i, 2 + j).value = round(float(Y[i, j]), 6)
            ws2.cell(2 + i, 2 + j).value = round(float(V[i, j]), 6)
    path = os.path.join(outdir, "result4.xlsx")
    wb.save(path)
    print("[write]", path, "(filled official template)")


def _write_paper_tables(outdir, X, Y, V):
    import csv
    tcols = [list(T_LIST).index(t) for t in PAPER_TIMES]
    with open(os.path.join(outdir, "paper_table_pos.csv"), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([""] + [f"{t} s" for t in PAPER_TIMES])
        for node, lab in zip(PAPER_NODES, PAPER_NODE_LABELS):
            w.writerow([f"{lab} x (m)"] + [f"{X[node,j]:.6f}" for j in tcols])
            w.writerow([f"{lab} y (m)"] + [f"{Y[node,j]:.6f}" for j in tcols])
    with open(os.path.join(outdir, "paper_table_speed.csv"), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([""] + [f"{t} s" for t in PAPER_TIMES])
        for node, lab in zip(PAPER_NODES, PAPER_NODE_LABELS):
            w.writerow([f"{lab} (m/s)"] + [f"{V[node,j]:.6f}" for j in tcols])
    print("[write] paper tables csv")


if __name__ == "__main__":
    main()
