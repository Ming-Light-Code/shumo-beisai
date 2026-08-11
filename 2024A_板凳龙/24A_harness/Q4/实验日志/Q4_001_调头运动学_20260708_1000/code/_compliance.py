"""result4.xlsx 合规自检（对照官方模板结构）+ 复制到项目根。"""
import os
import shutil
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(os.path.dirname(HERE), "outputs")
ROOT = HERE
for _ in range(9):
    if os.path.exists(os.path.join(ROOT, "题目")):
        break
    ROOT = os.path.dirname(ROOT)

tmpl = os.path.join(ROOT, "题目", "result4.xlsx")
mine = os.path.join(OUT, "result4.xlsx")
wt = openpyxl.load_workbook(tmpl)
wm = openpyxl.load_workbook(mine)

ok = True
print("sheet 名一致:", wt.sheetnames == wm.sheetnames, wm.sheetnames)
for s in wt.sheetnames:
    a, b = wt[s], wm[s]
    print(f"[{s}] 模板 {a.max_row}x{a.max_column} | 我 {b.max_row}x{b.max_column}",
          "OK" if (a.max_row == b.max_row and a.max_column == b.max_column) else "!!尺寸不符")
    if a.max_row != b.max_row or a.max_column != b.max_column:
        ok = False
    # 抽查表头/标签未被改动
    for r in range(1, a.max_row + 1):
        if a.cell(r, 1).value != b.cell(r, 1).value:
            print("  行标签改动 r=", r, repr(a.cell(r, 1).value), repr(b.cell(r, 1).value)); ok = False; break
    for c in range(1, a.max_column + 1):
        if a.cell(1, c).value != b.cell(1, c).value:
            print("  列标题改动 c=", c); ok = False; break

# 数据完整性：所有数据单元非空
ws1, ws2 = wm["位置"], wm["速度"]
empt = 0
for i in range(224):
    for j in range(201):
        if ws1.cell(2 + 2 * i, 2 + j).value is None: empt += 1
        if ws1.cell(3 + 2 * i, 2 + j).value is None: empt += 1
        if ws2.cell(2 + i, 2 + j).value is None: empt += 1
print("空数据单元数:", empt)
print("小数位抽查 位置(2,2)=", ws1.cell(2, 2).value, " 速度(2,2)=", ws2.cell(2, 2).value)
if empt: ok = False

print("==> 合规:", ok)
if ok:
    dst = os.path.join(ROOT, "result4.xlsx")
    shutil.copy(mine, dst)
    print("已复制到根:", dst)
