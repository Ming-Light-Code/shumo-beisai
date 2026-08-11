"""Q1 可视化：盘入快照 / 速度-节号曲线 / 龙头轨迹。依据 viz-standard skill。"""
import os
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt

# 字体：思源宋体未内置，回退 SimHei（图注标注）
FONT_NOTE = "字体回退 SimHei"
mpl.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
mpl.rcParams['font.family'] = 'sans-serif'
mpl.rcParams['axes.unicode_minus'] = False
PALETTE = ["#2E86C1", "#E74C3C", "#27AE60", "#F39C12", "#8E44AD", "#16A085", "#7F8C8D"]
mpl.rcParams['axes.prop_cycle'] = mpl.cycler(color=PALETTE)
mpl.rcParams.update({
    'font.size': 10, 'axes.titlesize': 14, 'axes.labelsize': 12,
    'xtick.labelsize': 10, 'ytick.labelsize': 10, 'legend.fontsize': 10,
    'figure.dpi': 150, 'savefig.dpi': 300, 'savefig.bbox': 'tight',
    'axes.grid': True, 'grid.alpha': 0.3,
})

P = 0.55
B = P / (2.0 * np.pi)
PAPER_TIMES = [0, 60, 120, 180, 240, 300]

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(os.path.dirname(HERE), "outputs")


def spiral_curve(theta_max):
    th = np.linspace(0, theta_max, 4000)
    r = B * th
    return r * np.cos(th), r * np.sin(th)


def fig1_snapshots():
    data = np.load(os.path.join(OUT, "snapshots.npz"))
    sx, sy = spiral_curve(20 * 2 * np.pi)
    fig, axes = plt.subplots(2, 3, figsize=(15, 10))
    for ax, t in zip(axes.ravel(), PAPER_TIMES):
        x = data[f"x_{t}"]
        y = data[f"y_{t}"]
        ax.plot(sx, sy, color="#B0B0B0", lw=0.6, alpha=0.7, label="盘入螺线")
        ax.plot(x, y, '-', color="#2E86C1", lw=1.0, label="板凳龙")
        ax.plot(x[0], y[0], '*', color="#E74C3C", ms=14, label="龙头前把手")
        ax.plot(x[-1], y[-1], 's', color="#27AE60", ms=6, label="龙尾后把手")
        ax.set_aspect('equal')
        ax.set_title(f"t = {t} s")
        ax.set_xlabel("x (m)")
        ax.set_ylabel("y (m)")
        ax.set_xlim(-13, 13)
        ax.set_ylim(-13, 13)
    axes[0, 0].legend(loc='upper right', fontsize=8)
    fig.suptitle(f"图1-1  板凳龙盘入过程快照（螺距0.55 m，龙头1 m/s；{FONT_NOTE}）", fontsize=14)
    p = os.path.join(OUT, "fig1-1_盘入快照.png")
    fig.savefig(p)
    fig.savefig(p.replace(".png", ".pdf"))
    plt.close(fig)
    print("[fig]", p)


def fig2_speed_profile():
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(OUT, "result1.xlsx"))
    ws = wb["速度"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    tcols = {t: header.index(f"{t} s") for t in PAPER_TIMES}
    speed = np.array([[r[c] for c in range(1, len(header))] for r in rows[1:]], dtype=float)
    node_idx = np.arange(speed.shape[0])
    fig, ax = plt.subplots(figsize=(10, 6))
    for t in PAPER_TIMES:
        ax.plot(node_idx, speed[:, tcols[t] - 1], lw=1.3, label=f"t = {t} s")
    ax.set_xlabel("把手节号 (0=龙头前, 223=龙尾后)")
    ax.set_ylabel("速度 (m/s)")
    ax.set_title("图1-2  各把手速度沿链分布")
    ax.legend()
    p = os.path.join(OUT, "fig1-2_速度沿链分布.png")
    fig.savefig(p)
    fig.savefig(p.replace(".png", ".pdf"))
    plt.close(fig)
    print("[fig]", p)


def fig3_head_track():
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(OUT, "result1.xlsx"))
    ws = wb["位置"]
    rows = list(ws.iter_rows(values_only=True))
    hx = np.array(rows[1][1:], dtype=float)  # 龙头x
    hy = np.array(rows[2][1:], dtype=float)  # 龙头y
    sx, sy = spiral_curve(33 * np.pi)
    fig, ax = plt.subplots(figsize=(8, 8))
    ax.plot(sx, sy, color="#B0B0B0", lw=0.6, alpha=0.7, label="盘入螺线")
    ax.plot(hx, hy, '-', color="#E74C3C", lw=1.6, label="龙头前把手轨迹 0–300 s")
    ax.plot(hx[0], hy[0], 'o', color="#2E86C1", ms=8, label="t=0 (A点)")
    ax.plot(hx[-1], hy[-1], '*', color="#27AE60", ms=14, label="t=300 s")
    ax.set_aspect('equal')
    ax.set_xlabel("x (m)")
    ax.set_ylabel("y (m)")
    ax.set_title("图1-3  龙头前把手 0–300 s 盘入轨迹")
    ax.legend()
    p = os.path.join(OUT, "fig1-3_龙头轨迹.png")
    fig.savefig(p)
    fig.savefig(p.replace(".png", ".pdf"))
    plt.close(fig)
    print("[fig]", p)


if __name__ == "__main__":
    fig1_snapshots()
    fig2_speed_profile()
    fig3_head_track()
    print("[done] figures ->", OUT)
