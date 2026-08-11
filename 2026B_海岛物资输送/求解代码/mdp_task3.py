"""
==============================================================================
 问题三：MDP (马尔可夫决策过程) 方案
==============================================================================

MDP形式化定义:
  状态 S: (x,y, O,H,F, M,Z, c, t, w)
    (x,y)    - 船舶位置
    O,H,F    - 自持资源 (燃油/淡水/食物)
    M        - 交易资金
    Z        - 目标物资
    c        - 当前作业点连续作业天数
    t        - 天数 (1..90)
    w        - 当日天气 (normal/storm), 行动前观测

  动作 A(s,w): {↑,↓,←,→, idle, work, buy(ΔO,ΔH,ΔF)}

  转移 P(s'|s,a,w):
    - 给定a和w, 资源消耗确定 (按天气消耗率)
    - 位置更新 (移动/idle/作业)
    - Z 在作业日增加对应作业点收益
    - 次日天气 w' ~ Bernoulli(0.8 normal, 0.2 storm), i.i.d.

  奖励 R(s): 仅终止状态非零
    R(到达E) = K * Z_T + M_T  (K >> Mmax, 实现字典序优先Z最大化)

求解方法: Approximate Dynamic Programming via Skeleton-Guided Rollout
  Level 1 (宏观): 骨架路径枚举 (B → {点集}* → E)
  Level 2 (微观): 对每条骨架, 定义天气感知策略并蒙特卡洛评估
  Level 3 (精化): 对最优骨架, 在关键决策点使用单步前瞻Rollout精化
==============================================================================
"""

import numpy as np
import random
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Tuple, Optional, Dict
import copy
import json
import time

# ============================================================================
# 问题参数 (Problem 3)
# ============================================================================
GRID_SIZE = 30

# 关键节点坐标
B = (1, 15)
E = (30, 15)
S1 = (12, 16)
S2 = (21, 16)
W1 = (6, 21)
W2 = (15, 9)
W3 = (24, 24)

# 工作点参数: (坐标, 日收益, 最大连续天数, 名称)
WORK_POINTS = [
    (W1, 20, 4, 'W1'),
    (W2, 15, 5, 'W2'),
    (W3, 28, 3, 'W3'),
]

# 补给平台
SUPPLY_POINTS = {'S1': S1, 'S2': S2}

# 初始资源
INIT_O, INIT_H, INIT_F = 100, 150, 100
INIT_M = 750
INIT_Z = 200
MAX_LOAD = 400
MAX_DAYS = 90

# 采购价格 (O, H, F)
PRICES = (2, 1, 2)

# 消耗率: [正常天气, 雷暴天气] → (O, H, F)
CONSUME_MOVE  = [(2, 3, 2), (8, 4, 3)]
CONSUME_IDLE  = [(1, 1, 1), (3, 3, 2)]
CONSUME_WORK  = [(5, 4, 3), (8, 6, 6)]

# 天气概率
P_NORMAL = 0.8
P_STORM  = 0.2

# 期望消耗率 (用于采购规划)
EXP_MOVE = (0.8*2 + 0.2*8, 0.8*3 + 0.2*4, 0.8*2 + 0.2*3)  # (3.2, 3.2, 2.2)
EXP_IDLE = (0.8*1 + 0.2*3, 0.8*1 + 0.2*3, 0.8*1 + 0.2*2)  # (1.4, 1.4, 1.2)
EXP_WORK = (0.8*5 + 0.2*8, 0.8*4 + 0.2*6, 0.8*3 + 0.2*6)  # (5.6, 4.4, 3.6)


def manhattan(p1, p2):
    return abs(p1[0] - p2[0]) + abs(p1[1] - p2[1])


def consume_rate(weather_idx, action_type):
    """weather_idx: 0=normal, 1=storm"""
    if action_type == 'move':
        return CONSUME_MOVE[weather_idx]
    elif action_type == 'idle':
        return CONSUME_IDLE[weather_idx]
    elif action_type == 'work':
        return CONSUME_WORK[weather_idx]
    return (0, 0, 0)


# ============================================================================
# 状态定义
# ============================================================================
@dataclass
class State:
    """MDP 状态 (不含天气, 天气在决策时观测)"""
    x: int
    y: int
    O: int
    H: int
    F: int
    M: int
    Z: int
    c: int = 0       # 连续作业天数
    day: int = 1     # 当前天数

    @property
    def pos(self):
        return (self.x, self.y)

    @pos.setter
    def pos(self, value):
        self.x, self.y = value

    def clone(self):
        return copy.deepcopy(self)

    def is_feasible(self):
        if self.O < 0 or self.H < 0 or self.F < 0 or self.M < 0:
            return False
        if self.O + self.H + self.F > MAX_LOAD:
            return False
        return True

    def total_load(self):
        return self.O + self.H + self.F


# ============================================================================
# 骨架枚举器
# ============================================================================
class SkeletonEnumerator:
    """枚举从B到E的可行骨架路径"""

    def __init__(self):
        self.nodes = {
            'B': B, 'E': E,
            'W1': W1, 'W2': W2, 'W3': W3,
            'S1': S1, 'S2': S2,
        }
        self.work_names = {'W1', 'W2', 'W3'}
        self.supply_names = {'S1', 'S2'}
        self.intermediate = ['W1', 'W2', 'W3', 'S1', 'S2']

        # 预计算所有节点间距离
        self.dist = {}
        names = list(self.nodes.keys())
        for n1 in names:
            for n2 in names:
                self.dist[(n1, n2)] = manhattan(self.nodes[n1], self.nodes[n2])

    def enumerate(self, max_intermediate=8, max_skeletons=5000):
        """枚举骨架路径, 最多 max_intermediate 个中间节点"""
        skeletons = []

        def dfs(current, path, total_travel):
            nonlocal skeletons
            if len(skeletons) >= max_skeletons:
                return

            # 尝试直接去E
            dE = self.dist[(current, 'E')]
            if total_travel + dE <= MAX_DAYS:
                skeletons.append(tuple(path + ['E']))

            # 尝试中间节点
            for nxt in self.intermediate:
                if nxt == current:
                    continue
                if current in self.supply_names and nxt in self.supply_names:
                    continue
                if len(path) > max_intermediate:
                    continue

                d = self.dist[(current, nxt)]
                if d == 0 or total_travel + d > MAX_DAYS:
                    continue

                dfs(nxt, path + [nxt], total_travel + d)

        dfs('B', ['B'], 0)
        return skeletons

    def skeleton_distance(self, skeleton):
        """计算骨架的总路径长度"""
        total = 0
        for i in range(len(skeleton) - 1):
            total += self.dist[(skeleton[i], skeleton[i+1])]
        return total

    def skeleton_info(self, skeleton):
        """提取骨架信息"""
        work_nodes = []
        supply_nodes = []
        for name in skeleton[1:-1]:
            if name in self.work_names:
                work_nodes.append(name)
            elif name in self.supply_names:
                supply_nodes.append(name)
        return {
            'skeleton': skeleton,
            'travel_days': self.skeleton_distance(skeleton),
            'work_nodes': work_nodes,
            'supply_nodes': supply_nodes,
            'max_work_days': sum(
                {'W1': 4, 'W2': 5, 'W3': 3}[w] for w in work_nodes
            ),
        }


# ============================================================================
# MDP 仿真器 (骨架引导策略)
# ============================================================================
class SkeletonSimulator:
    """对给定骨架路径, 模拟一次天气序列下的航行"""

    def __init__(self):
        self.node_pos = {'B': B, 'E': E, 'W1': W1, 'W2': W2, 'W3': W3,
                         'S1': S1, 'S2': S2}
        self.work_gains = {'W1': 20, 'W2': 15, 'W3': 28}
        self.work_max = {'W1': 4, 'W2': 5, 'W3': 3}

    @staticmethod
    def sample_weather():
        return 0 if random.random() < P_NORMAL else 1

    def simulate(self, skeleton) -> Optional[Tuple[int, int, List[Dict]]]:
        """模拟一次航行, 返回 (Z_final, M_final, daily_log) 或 None"""
        state = State(x=B[0], y=B[1], O=INIT_O, H=INIT_H, F=INIT_F,
                       M=INIT_M, Z=INIT_Z, c=0, day=1)
        log = []
        seg_idx = 1

        while state.day <= MAX_DAYS:
            if state.pos == E:
                return state.Z, state.M, log

            target_name = skeleton[seg_idx]
            target_pos = self.node_pos[target_name]

            if state.pos == target_pos:
                node_name = target_name
                if node_name in self.work_gains:
                    weather = self.sample_weather()
                    if weather == 0 and state.c < self.work_max[node_name]:
                        action = 'work'
                        ns = self._apply_action(state, action, weather,
                                                 work_gain=self.work_gains[node_name])
                        if ns is None: return None
                        log.append(self._mk_log(state, action, weather, node_name, None))
                        state = ns
                    else:
                        action = 'idle'
                        ns = self._apply_action(state, action, weather)
                        if ns is None: return None
                        ns.c = 0
                        log.append(self._mk_log(state, action, weather, node_name, None))
                        state = ns
                        seg_idx += 1
                        if seg_idx >= len(skeleton): return None
                elif node_name in ['S1', 'S2']:
                    weather = self.sample_weather()
                    buy_O, buy_H, buy_F = self._compute_buy(state, skeleton, seg_idx)
                    ns = self._apply_buy(state, buy_O, buy_H, buy_F)
                    if ns is None: return None
                    log.append(self._mk_log(state, f'buy({buy_O},{buy_H},{buy_F})',
                                             weather, node_name, (buy_O,buy_H,buy_F)))
                    ns2 = self._apply_action(ns, 'idle', weather)
                    if ns2 is None: return None
                    log.append(self._mk_log(ns, 'idle', weather, node_name, None))
                    state = ns2
                    seg_idx += 1
                    if seg_idx >= len(skeleton): return None
                else:
                    seg_idx += 1
                    continue
            else:
                weather = self.sample_weather()
                dx = np.sign(target_pos[0] - state.x) or 0
                dy = np.sign(target_pos[1] - state.y) or 0
                if dx == 0 and dy == 0:
                    seg_idx += 1; continue
                new_pos = (state.x + dx, state.y + dy)
                ns = self._apply_action(state, 'move', weather, new_pos=new_pos)
                if ns is None: return None
                log.append(self._mk_log(state, 'move', weather, None, None, new_pos=new_pos))
                state = ns
        return None

    def _apply_action(self, state, action, weather, new_pos=None, work_gain=0):
        s = state.clone()
        cr = consume_rate(weather, action)
        s.O -= cr[0]; s.H -= cr[1]; s.F -= cr[2]
        if action == 'work': s.Z += work_gain; s.c += 1
        elif action == 'move': s.pos = new_pos; s.c = 0
        elif action == 'idle': s.c = 0
        s.day += 1
        return s if s.is_feasible() else None

    def _apply_buy(self, state, buy_O, buy_H, buy_F):
        s = state.clone()
        cost = buy_O * PRICES[0] + buy_H * PRICES[1] + buy_F * PRICES[2]
        if cost > s.M: return None
        s.O += buy_O; s.H += buy_H; s.F += buy_F; s.M -= cost
        return s if s.is_feasible() else None

    def _compute_buy(self, state, skeleton, seg_idx):
        """在供给点使用期望消耗率 + 安全余量计算最优采购量"""
        need_O, need_H, need_F = 0, 0, 0
        next_supply_idx = len(skeleton) - 1
        for j in range(seg_idx + 1, len(skeleton)):
            if skeleton[j] in ['S1', 'S2']:
                next_supply_idx = j; break
        for j in range(seg_idx, next_supply_idx):
            n1, n2 = skeleton[j], skeleton[j+1]
            d = manhattan(self.node_pos[n1], self.node_pos[n2])
            need_O += EXP_MOVE[0]*d; need_H += EXP_MOVE[1]*d; need_F += EXP_MOVE[2]*d
            if n2 in self.work_gains:
                exp_w = int(P_NORMAL * self.work_max[n2])
                need_O += EXP_WORK[0]*exp_w; need_H += EXP_WORK[1]*exp_w; need_F += EXP_WORK[2]*exp_w
        safety = 1.2
        need_O = int(np.ceil(need_O*safety))
        need_H = int(np.ceil(need_H*safety))
        need_F = int(np.ceil(need_F*safety))
        buy_O = max(0, need_O - state.O)
        buy_H = max(0, need_H - state.H)
        buy_F = max(0, need_F - state.F)
        space = MAX_LOAD - state.total_load()
        total_buy = buy_O + buy_H + buy_F
        if total_buy > space and total_buy > 0:
            scale = space / total_buy
            buy_O = int(buy_O*scale); buy_H = int(buy_H*scale); buy_F = int(buy_F*scale)
        cost = buy_O*PRICES[0] + buy_H*PRICES[1] + buy_F*PRICES[2]
        if cost > state.M and cost > 0:
            scale = state.M / cost
            buy_O = int(buy_O*scale); buy_H = int(buy_H*scale); buy_F = int(buy_F*scale)
        return buy_O, buy_H, buy_F

    def _mk_log(self, state, action, weather, node_name, buy_amounts, new_pos=None):
        wx = 'N' if weather == 0 else 'S'
        return {'day': state.day, 'pos': (state.x, state.y),
                'O': state.O, 'H': state.H, 'F': state.F,
                'M': state.M, 'Z': state.Z, 'c': state.c,
                'weather': wx, 'action': action, 'node': node_name,
                'new_pos': new_pos, 'buy': buy_amounts}


# ============================================================================
# MDP Rollout 评估器
# ============================================================================
class MDPEvaluator:
    """对骨架路径进行蒙特卡洛评估"""
    def __init__(self, simulator):
        self.sim = simulator

    def evaluate(self, skeleton, n_simulations=200):
        total_Z, total_M, successes = 0, 0, 0
        best_Z, best_M = -1, -1
        best_log = None
        for _ in range(n_simulations):
            result = self.sim.simulate(skeleton)
            if result is not None:
                Z_f, M_f, log = result
                total_Z += Z_f; total_M += M_f; successes += 1
                if Z_f > best_Z or (Z_f == best_Z and M_f > best_M):
                    best_Z, best_M = Z_f, M_f; best_log = log
        if successes == 0: return None
        return {'skeleton': skeleton, 'avg_Z': total_Z/successes,
                'avg_M': total_M/successes, 'success_rate': successes/n_simulations,
                'best_Z': best_Z, 'best_M': best_M, 'best_log': best_log,
                'n_successes': successes}


# ============================================================================
# MDP Rollout 精化器
# ============================================================================
class MDPRolloutRefiner:
    """在工作点使用单步前瞻 Rollout 精化决策"""
    def __init__(self, simulator, base_skeleton):
        self.sim = simulator; self.base_skeleton = base_skeleton
        self.work_gains = {'W1': 20, 'W2': 15, 'W3': 28}
        self.work_max = {'W1': 4, 'W2': 5, 'W3': 3}
        self.node_pos = {'B': B, 'E': E, 'W1': W1, 'W2': W2, 'W3': W3,
                         'S1': S1, 'S2': S2}

    def decide(self, state, weather, work_node, skeleton, seg_idx, n_rollouts=30):
        actions = []
        if weather == 0 and state.c < self.work_max[work_node]:
            actions.append('work')
        actions.append('idle')
        if len(actions) == 1: return actions[0]
        best_a, best_v = None, -float('inf')
        for a in actions:
            v = self._q_est(state, a, weather, work_node, skeleton, seg_idx, n_rollouts)
            if v > best_v: best_v = v; best_a = a
        return best_a

    def _q_est(self, state, action, weather, work_node, skeleton, seg_idx, n):
        total_Z, total_M, cnt = 0, 0, 0
        for _ in range(n):
            s = state.clone()
            if action == 'work':
                cr = consume_rate(weather, 'work')
                s.O -= cr[0]; s.H -= cr[1]; s.F -= cr[2]
                s.Z += self.work_gains[work_node]; s.c += 1; s.day += 1
                if not s.is_feasible(): continue
                ns = seg_idx
            else:
                cr = consume_rate(weather, 'idle')
                s.O -= cr[0]; s.H -= cr[1]; s.F -= cr[2]
                s.c = 0; s.day += 1
                if not s.is_feasible(): continue
                ns = seg_idx + 1
            result = self._continue(s, skeleton, ns)
            if result: Z_f, M_f, _ = result; total_Z += Z_f; total_M += M_f; cnt += 1
        return -float('inf') if cnt == 0 else total_Z/cnt*100000 + total_M/cnt

    def _continue(self, state, skeleton, seg_idx):
        work_max = self.work_max; work_gains = self.work_gains
        while state.day <= MAX_DAYS:
            if state.pos == E: return state.Z, state.M, []
            if seg_idx >= len(skeleton): return None
            tgt = skeleton[seg_idx]; tpos = self.node_pos[tgt]
            if state.pos == tpos:
                if tgt in work_gains:
                    w = self.sim.sample_weather()
                    if w == 0 and state.c < work_max[tgt]:
                        state = self._aa(state, 'work', w, wg=work_gains[tgt])
                    else:
                        state = self._aa(state, 'idle', w); state.c = 0; seg_idx += 1
                    if state is None: return None
                elif tgt in ['S1','S2']:
                    w = self.sim.sample_weather()
                    buy = self.sim._compute_buy(state, skeleton, seg_idx)
                    state = self._ab(state, *buy)
                    if state is None: return None
                    state = self._aa(state, 'idle', w)
                    if state is None: return None
                    seg_idx += 1
                else: seg_idx += 1
            else:
                w = self.sim.sample_weather()
                dx = np.sign(tpos[0]-state.x) or 0
                dy = np.sign(tpos[1]-state.y) or 0
                if dx==0 and dy==0: seg_idx+=1; continue
                state = self._aa(state, 'move', w, np=(state.x+dx, state.y+dy))
                if state is None: return None
        return None

    def _aa(self, s, a, w, np=None, wg=0):
        s = s.clone(); cr = consume_rate(w, a)
        s.O-=cr[0]; s.H-=cr[1]; s.F-=cr[2]
        if a=='work': s.Z+=wg; s.c+=1
        elif a=='move': s.pos=np; s.c=0
        else: s.c=0
        s.day+=1; return s if s.is_feasible() else None

    def _ab(self, s, bo, bh, bf):
        s=s.clone(); c=bo*2+bh*1+bf*2
        if c>s.M: return None
        s.O+=bo; s.H+=bh; s.F+=bf; s.M-=c
        return s if s.is_feasible() else None


# ============================================================================
# MDP 主求解器
# ============================================================================
class MDPSolver:
    def __init__(self, seed=42):
        self.seed = seed
        random.seed(seed); np.random.seed(seed)
        self.enumerator = SkeletonEnumerator()
        self.simulator = SkeletonSimulator()
        self.evaluator = MDPEvaluator(self.simulator)
        self.results = []; self.best_result = None

    def solve(self, n_simulations=500, top_k=5, refine=True):
        t0 = time.time()
        print("="*70)
        print("Phase 1: 骨架枚举"); print("="*70)
        skeletons = self.enumerator.enumerate(max_intermediate=8, max_skeletons=2000)
        print(f"  生成 {len(skeletons)} 条候选骨架")
        unique = list(set(skeletons))
        unique.sort(key=lambda s: self.enumerator.skeleton_distance(s))
        work_names = {'W1','W2','W3'}
        sw = [s for s in unique if any(n in work_names for n in s[1:-1])]
        all_candidates = sw + [('B','E')]
        print(f"  去重 {len(unique)}, 含工作点 {len(sw)}, 总计 {len(all_candidates)}")

        print("\n"+"="*70)
        print("Phase 2: MDP 蒙特卡洛评估"); print("="*70)
        results = []
        for i, skel in enumerate(all_candidates):
            if i%200==0:
                print(f"  进度: {i}/{len(all_candidates)} (耗时 {time.time()-t0:.1f}s)")
            info = self.enumerator.skeleton_info(skel)
            if info['travel_days'] > 80: continue
            r = self.evaluator.evaluate(skel, n_simulations=n_simulations)
            if r: results.append(r)
        results.sort(key=lambda r: (r['avg_Z'], r['avg_M']), reverse=True)
        self.results = results
        print(f"\n  可行骨架 {len(results)} 条, 耗时 {time.time()-t0:.1f}s")

        print("\n"+"="*70)
        print("Phase 3: Top 结果"); print("="*70)
        for rank, r in enumerate(results[:top_k]):
            info = self.enumerator.skeleton_info(r['skeleton'])
            print(f"  #{rank+1}: {' → '.join(r['skeleton'])}")
            print(f"       travel={info['travel_days']}d  avg_Z={r['avg_Z']:.1f}  "
                  f"avg_M={r['avg_M']:.1f}  成功率={r['n_successes']}/{n_simulations}  "
                  f"best_Z={r['best_Z']} best_M={r['best_M']}")

        self.best_result = results[0]
        print(f"\n  最优骨架: {' → '.join(self.best_result['skeleton'])}")
        print(f"  期望 Z={self.best_result['avg_Z']:.1f}, M={self.best_result['avg_M']:.1f}")

        if refine and self.best_result:
            print("\n"+"="*70)
            print("Phase 4: MDP Rollout 精化"); print("="*70)
            refined = self._refine(n_simulations=300)
            print(f"  精化后: Z={refined['avg_Z']:.1f}, M={refined['avg_M']:.1f}")
            if refined['avg_Z'] >= self.best_result['avg_Z']:
                self.best_result = refined

        print(f"\n  总耗时: {time.time()-t0:.1f}s")
        return self.best_result

    def _refine(self, n_simulations=300):
        best = self.best_result; skel = best['skeleton']
        refiner = MDPRolloutRefiner(self.simulator, skel)
        node_pos = {'B':B,'E':E,'W1':W1,'W2':W2,'W3':W3,'S1':S1,'S2':S2}
        wg = {'W1':20,'W2':15,'W3':28}
        total_Z, total_M, succ = 0, 0, 0
        bZ, bM = -1, -1; bLog = None
        for _ in range(n_simulations):
            s = State(x=B[0],y=B[1],O=INIT_O,H=INIT_H,F=INIT_F,M=INIT_M,Z=INIT_Z,c=0,day=1)
            log, seg, fail = [], 1, False
            while s.day <= MAX_DAYS:
                if s.pos == E:
                    total_Z+=s.Z; total_M+=s.M; succ+=1
                    if s.Z>bZ or (s.Z==bZ and s.M>bM): bZ,bM,bLog = s.Z,s.M,log
                    break
                if seg>=len(skel): fail=True; break
                tn = skel[seg]; tp = node_pos[tn]
                if s.pos == tp:
                    if tn in wg:
                        w = self.simulator.sample_weather()
                        a = refiner.decide(s, w, tn, skel, seg, n_rollouts=30)
                        if a=='work':
                            cr=consume_rate(w,'work'); s.O-=cr[0];s.H-=cr[1];s.F-=cr[2]
                            s.Z+=wg[tn]; s.c+=1; s.day+=1
                            if not s.is_feasible(): fail=True; break
                        else:
                            cr=consume_rate(w,'idle'); s.O-=cr[0];s.H-=cr[1];s.F-=cr[2]
                            s.c=0; s.day+=1
                            if not s.is_feasible(): fail=True; break
                            seg+=1
                    elif tn in ['S1','S2']:
                        w=self.simulator.sample_weather()
                        buy=self.simulator._compute_buy(s,skel,seg)
                        if buy[0]*2+buy[1]*1+buy[2]*2 > s.M: fail=True; break
                        s.O+=buy[0];s.H+=buy[1];s.F+=buy[2];s.M-=buy[0]*2+buy[1]*1+buy[2]*2
                        if not s.is_feasible(): fail=True; break
                        cr=consume_rate(w,'idle'); s.O-=cr[0];s.H-=cr[1];s.F-=cr[2];s.day+=1
                        if not s.is_feasible(): fail=True; break
                        seg+=1
                    else: seg+=1
                else:
                    w=self.simulator.sample_weather()
                    dx=np.sign(tp[0]-s.x) or 0; dy=np.sign(tp[1]-s.y) or 0
                    if dx==0 and dy==0: seg+=1; continue
                    cr=consume_rate(w,'move'); s.O-=cr[0];s.H-=cr[1];s.F-=cr[2]
                    s.pos=(s.x+dx,s.y+dy); s.c=0; s.day+=1
                    if not s.is_feasible(): fail=True; break
            if fail: continue
        return {'skeleton':skel,'avg_Z':total_Z/succ if succ>0 else 0,
                'avg_M':total_M/succ if succ>0 else 0,
                'success_rate':succ/n_simulations,'best_Z':bZ,'best_M':bM,
                'best_log':bLog,'n_successes':succ}


# ============================================================================
# 结果输出到 Excel
# ============================================================================
def generate_daily_schedule(best_result, output_path='result.xls'):
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.run(['pip','install','openpyxl'], check=True)
        import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "MDP方案"
    headers = ['天数','位置X','位置Y','天气','行为','工作点',
               '燃油O','淡水H','食物F','资金M','目标物资Z','连续作业c',
               '采购O','采购H','采购F']
    for c,h in enumerate(headers,1): ws.cell(row=1,column=c,value=h)
    log = best_result.get('best_log',[])
    for c,v in enumerate([0,B[0],B[1],'-','初始','B',INIT_O,INIT_H,INIT_F,INIT_M,INIT_Z,0,0,0,0],1):
        ws.cell(row=2,column=c,value=v)
    for i,e in enumerate(log,3):
        ws.cell(row=i,column=1,value=e['day'])
        ws.cell(row=i,column=2,value=e['pos'][0]); ws.cell(row=i,column=3,value=e['pos'][1])
        ws.cell(row=i,column=4,value=e['weather']); ws.cell(row=i,column=5,value=e['action'])
        ws.cell(row=i,column=6,value=e.get('node','') or '')
        ws.cell(row=i,column=7,value=e['O']); ws.cell(row=i,column=8,value=e['H'])
        ws.cell(row=i,column=9,value=e['F']); ws.cell(row=i,column=10,value=e['M'])
        ws.cell(row=i,column=11,value=e['Z']); ws.cell(row=i,column=12,value=e['c'])
        buy = e.get('buy')
        ws.cell(row=i,column=13,value=buy[0] if buy else 0)
        ws.cell(row=i,column=14,value=buy[1] if buy else 0)
        ws.cell(row=i,column=15,value=buy[2] if buy else 0)
    wb.save(output_path)
    print(f"\n  每日排程已输出至: {output_path}")


# ============================================================================
# 主程序
# ============================================================================
if __name__ == '__main__':
    print("="*70)
    print("  问题三 · MDP (马尔可夫决策过程) 求解")
    print("  天气随机先验: P(正常)=0.8, P(雷暴)=0.2, 每日独立")
    print("="*70)
    solver = MDPSolver(seed=42)
    best = solver.solve(n_simulations=500, top_k=5, refine=True)
    if best and best.get('best_log'):
        op = r'C:\Users\ming\Desktop\数模备赛\result.xls'
        generate_daily_schedule(best, op)
    result_json = {
        'problem': 'Task 3 (MDP)',
        'parameters': {
            'grid':'30x30','B':'(1,15)','E':'(30,15)',
            'S1':'(12,16)','S2':'(21,16)',
            'W1':'(6,21) gain=20 max=4','W2':'(15,9) gain=15 max=5',
            'W3':'(24,24) gain=28 max=3',
            'P_normal':0.8,'P_storm':0.2,'max_days':90,'max_load':400,
            'init_O':100,'init_H':150,'init_F':100,'init_M':750,'init_Z':200
        },
        'best_skeleton': list(best['skeleton']),
        'expected_Z': best['avg_Z'], 'expected_M': best['avg_M'],
        'success_rate': best['success_rate'],
        'best_episode_Z': best['best_Z'], 'best_episode_M': best['best_M']
    }
    jp = r'C:\Users\ming\Desktop\数模备赛\mdp_task3_result.json'
    with open(jp, 'w', encoding='utf-8') as f:
        json.dump(result_json, f, ensure_ascii=False, indent=2)
    print(f"\n  结果JSON: {jp}")
    print("\n"+"="*70)
    print("  求解完成!")
    print("="*70)
