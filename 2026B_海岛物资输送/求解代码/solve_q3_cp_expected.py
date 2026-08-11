# -*- coding: utf-8 -*-
"""
2026B 问题三 CP 期望值求解器
方法：约束规划 + 期望值消耗参数 + 差异化安全裕度
策略：路径骨架 DFS 枚举 + 工作天数枚举 + 停泊重置 + 上界剪枝
"""
import math
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ========== 全局参数 ==========
MAX_DAYS  = 90; MAX_LOAD  = 400
INIT_O = 100; INIT_H = 150; INIT_F = 100
INIT_M = 750; INIT_Z = 200
PN = 0.8; PS = 0.2
PRICE_O = 2; PRICE_H = 1; PRICE_F = 2
SAFETY_O = 1.25; SAFETY_H = 1.05; SAFETY_F = 1.05

class Cons: pass
cons_exp = Cons()
cons_exp.MO = PN*2 + PS*8; cons_exp.MH = PN*3 + PS*4; cons_exp.MF = PN*2 + PS*3
cons_exp.PO = PN*1 + PS*3; cons_exp.PH = PN*1 + PS*3; cons_exp.PF = PN*1 + PS*2
cons_exp.WO = PN*5 + PS*8; cons_exp.WH = PN*4 + PS*6; cons_exp.WF = PN*3 + PS*6

cons_plan = Cons()
cons_plan.MO = cons_exp.MO*SAFETY_O; cons_plan.MH = cons_exp.MH*SAFETY_H; cons_plan.MF = cons_exp.MF*SAFETY_F
cons_plan.PO = cons_exp.PO*SAFETY_O; cons_plan.PH = cons_exp.PH*SAFETY_H; cons_plan.PF = cons_exp.PF*SAFETY_F
cons_plan.WO = cons_exp.WO*SAFETY_O; cons_plan.WH = cons_exp.WH*SAFETY_H; cons_plan.WF = cons_exp.WF*SAFETY_F

POINTS = {"B":(1,15),"E":(30,15),"W1":(6,21),"W2":(15,9),"W3":(24,24),"S1":(12,16),"S2":(21,16)}
WORK_INFO = {"W1":(20,4),"W2":(15,5),"W3":(28,3)}
INTER_NODES = ["W1","W2","W3","S1","S2"]
ALL_NAMES = ["B","E","W1","W2","W3","S1","S2"]

def manhattan(a,b): return abs(a[0]-b[0])+abs(a[1]-b[1])
D = [[manhattan(POINTS[ALL_NAMES[i]],POINTS[ALL_NAMES[j]]) for j in range(7)] for i in range(7)]
def dist(a,b): return D[ALL_NAMES.index(a)][ALL_NAMES.index(b)]

def max_work_with_park(wm, R):
    if R<=0 or wm<=0: return 0
    best=0
    for k in range(1,R+2):
        need=k*wm+(k-1)
        if need<=R: best=max(best,k*wm+min(wm,R-need))
        else: break
    return max(best,min(wm,R))

def enum_skeletons(max_inter=8, max_cnt=50000):
    skels=[]
    def dfs(path,td):
        if len(skels)>=max_cnt: return
        cur=path[-1]; dE=dist(cur,"E")
        if td+dE<=MAX_DAYS: skels.append(path+["E"])
        if len(path)-1>=max_inter: return
        for nxt in INTER_NODES:
            if nxt==cur: continue
            d=dist(cur,nxt)
            if d==0 or td+d+dist(nxt,"E")>MAX_DAYS: continue
            dfs(path+[nxt],td+d)
    dfs(["B"],0)
    return skels

def supply_needs(st,sk,seg):
    ns=len(sk)
    for j in range(seg+1,len(sk)):
        if sk[j] in ("S1","S2","E"): ns=j; break
    no=nh=nf=0.0
    for j in range(seg,min(ns,len(sk)-1)):
        a,b=sk[j],sk[j+1]; d=dist(a,b)
        no+=cons_plan.MO*d; nh+=cons_plan.MH*d; nf+=cons_plan.MF*d
        if b in WORK_INFO:
            _,wm=WORK_INFO[b]
            no+=cons_plan.WO*wm+cons_plan.PO*max(0,wm//3)
            nh+=cons_plan.WH*wm+cons_plan.PH*max(0,wm//3)
            nf+=cons_plan.WF*wm+cons_plan.PF*max(0,wm//3)
    bo=max(0,int(math.ceil(no-st["O"])))
    bh=max(0,int(math.ceil(nh-st["H"])))
    bf=max(0,int(math.ceil(nf-st["F"])))
    sp=MAX_LOAD-(st["O"]+st["H"]+st["F"])
    tot=bo+bh+bf
    if tot>sp and sp>0: sc=sp/tot; bo=int(bo*sc); bh=int(bh*sc); bf=int(bf*sc)
    cst=bo*PRICE_O+bh*PRICE_H+bf*PRICE_F
    if cst>st["M"] and st["M"]>0: sc=st["M"]/cst; bo=int(bo*sc); bh=int(bh*sc); bf=int(bf*sc)
    return bo,bh,bf

def simulate(sk,wd_list,pk_list):
    O,H,F,M,Z=INIT_O,INIT_H,INIT_F,INIT_M,INIT_Z
    day=1; log=[]
    log.append(dict(day=0,x=POINTS["B"][0],y=POINTS["B"][1],weather="-",action="Start",
        point="B",O=O,H=H,F=F,M=M,Z=Z,consec=0))
    wi=0
    for seg in range(len(sk)-1):
        cur=sk[seg]; nxt=sk[seg+1]; d=dist(cur,nxt)
        pk=pk_list[seg] if seg<len(pk_list) else 0
        for _ in range(pk):
            if day>MAX_DAYS: return False,Z,M,log
            O-=cons_exp.PO; H-=cons_exp.PH; F-=cons_exp.PF
            log.append(dict(day=day,x=POINTS[cur][0],y=POINTS[cur][1],weather="N",
                action="park(at sea)",point="",O=max(0,O),H=max(0,H),F=max(0,F),M=M,Z=Z,consec=0))
            if O<0 or H<0 or F<0: return False,Z,M,log
            day+=1
        cx,cy=POINTS[cur]; tx,ty=POINTS[nxt]
        for _ in range(d):
            if day>MAX_DAYS: return False,Z,M,log
            if cx!=tx: cx+=1 if tx>cx else -1
            elif cy!=ty: cy+=1 if ty>cy else -1
            O-=cons_exp.MO; H-=cons_exp.MH; F-=cons_exp.MF
            log.append(dict(day=day,x=cx,y=cy,weather="N",action="move",point="",
                O=max(0,O),H=max(0,H),F=max(0,F),M=M,Z=Z,consec=0))
            if O<0 or H<0 or F<0: return False,Z,M,log
            day+=1
        if nxt in ("S1","S2"):
            bo,bh,bf=supply_needs(dict(O=O,H=H,F=F,M=M),sk,seg+1)
            cst=bo*PRICE_O+bh*PRICE_H+bf*PRICE_F
            if cst<=M: O+=bo; H+=bh; F+=bf; M-=cst
            O-=cons_exp.PO; H-=cons_exp.PH; F-=cons_exp.PF
            log.append(dict(day=day,x=POINTS[nxt][0],y=POINTS[nxt][1],weather="N",
                action="buy(%d,%d,%d)"%(bo,bh,bf),point=nxt,O=max(0,O),H=max(0,H),F=max(0,F),M=M,Z=Z,consec=0))
            if O<0 or H<0 or F<0 or M<0: return False,Z,M,log
            day+=1
        elif nxt in WORK_INFO:
            if wi>=len(wd_list): continue
            wd=wd_list[wi]; wi+=1
            if wd==0: continue
            gain,wm=WORK_INFO[nxt]; consec=0; done=0
            while done<wd:
                if day>MAX_DAYS: return False,Z,M,log
                if consec<wm:
                    O-=cons_exp.WO; H-=cons_exp.WH; F-=cons_exp.WF; Z+=gain; consec+=1; done+=1
                    log.append(dict(day=day,x=POINTS[nxt][0],y=POINTS[nxt][1],weather="N",
                        action="work",point=nxt,O=max(0,O),H=max(0,H),F=max(0,F),M=M,Z=Z,consec=consec))
                else:
                    O-=cons_exp.PO; H-=cons_exp.PH; F-=cons_exp.PF; consec=0
                    log.append(dict(day=day,x=POINTS[nxt][0],y=POINTS[nxt][1],weather="N",
                        action="park(reset)",point=nxt,O=max(0,O),H=max(0,H),F=max(0,F),M=M,Z=Z,consec=0))
                if O<0 or H<0 or F<0: return False,Z,M,log
                day+=1
        elif nxt=="E":
            return True,Z,M,log
    return False,Z,M,log

def cp_search():
    print("=== CP Expectation-Value Solver for Task 3 ===\n")
    print("[Phase 1] Enumerating skeletons...")
    skels=enum_skeletons()
    seen=set(); uniq=[]
    for sk in skels:
        k="|".join(sk)
        if k not in seen: seen.add(k); uniq.append(sk)
    uniq.sort(key=lambda s:sum(dist(s[i],s[i+1]) for i in range(len(s)-1)))
    hw=[sk for sk in uniq if any(n in WORK_INFO for n in sk)]
    print("  Raw: %d, Unique: %d, With work: %d"%(len(skels),len(uniq),len(hw)))
    print("\n[Phase 2] CP search...")
    bestZ=-1; bestM=-1; bestSk=None; bestWd=None; bestPk=None; bestLog=None
    explored=0; pruned=0
    for idx,sk in enumerate(hw):
        wps=[(i,n) for i,n in enumerate(sk) if n in WORK_INFO]
        if not wps: continue
        nwp=len(wps); td=sum(dist(sk[i],sk[i+1]) for i in range(len(sk)-1))
        rem=MAX_DAYS-td
        ubZ=INIT_Z+max_work_with_park(3,rem)*28
        if ubZ<=bestZ: pruned+=1
        elif ubZ>bestZ:
            max_wd=[max_work_with_park(WORK_INFO[n][1],rem) for _,n in wps]
            ranges=[]
            for mw in max_wd:
                if mw<=3: ranges.append(list(range(mw+1)))
                else: ranges.append([0,mw//3,2*mw//3,mw])
            nlegs=len(sk)-1
            for wdc in product(*ranges):
                twd=sum(wdc); epk=0
                for wi2,(_,n) in enumerate(wps):
                    if wdc[wi2]>0: epk+=max(0,(wdc[wi2]+WORK_INFO[n][1]-1)//WORK_INFO[n][1]-1)
                if td+twd+epk>MAX_DAYS: continue
                prem=MAX_DAYS-td-twd-epk
                if prem<=0: pc=[(0,)*nlegs]
                else: pc=[]; [pc.append((p,)+(0,)*(nlegs-1)) for p in range(min(prem+1,6))]
                for pk in pc:
                    explored+=1
                    ok,Zf,Mf,log=simulate(sk,list(wdc),list(pk))
                    if ok and (Zf>bestZ or (Zf==bestZ and Mf>bestM)):
                        bestZ,bestM,bestSk,bestWd,bestPk,bestLog=Zf,Mf,sk,list(wdc),list(pk),log
        if idx%2000==0:
            print("  [%d/%d] best:Z=%d M=%d pruned=%d explored=%d"%(idx,len(hw),bestZ,bestM,pruned,explored))
    print("\n  Total: explored=%d pruned=%d"%(explored,pruned))
    print("\n"+"="*60)
    print("  OPTIMAL SOLUTION")
    print("="*60)
    if bestSk:
        print("  Path: %s"%" -> ".join(bestSk))
        print("  Z = %d"%(int(bestZ)))
        print("  M = %d"%(int(bestM)))
        wns=[n for n in bestSk if n in WORK_INFO]
        if bestWd: print("  Work: %s"%dict(zip(wns,bestWd)))
        if bestPk: print("  Park: %s"%bestPk)
        td2=len([e for e in bestLog if e["day"]>0])
        print("  Days: %d/%d"%(td2,MAX_DAYS))
    return bestSk,bestWd,bestPk,bestZ,bestM,bestLog

def export_xlsx(sk,wd,pk,Z,M,log,fn):
    wb=Workbook(); ws=wb.active; ws.title="Task3 CP Solution"
    hd=["Day","PosX","PosY","Weather","Action","Point","FuelO","WaterH","FoodF","MoneyM","TargetZ","Consec","BuyO","BuyH","BuyF"]
    hf=Font(bold=True,size=10,color="FFFFFF"); hfl=PatternFill(start_color="4472C4",end_color="4472C4",fill_type="solid")
    bd=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    for c,h in enumerate(hd,1):
        cl=ws.cell(row=1,column=c,value=h); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal="center"); cl.border=bd
    for r,e in enumerate(log,2):
        for c,k in enumerate(["day","x","y","weather","action","point","O","H","F","M","Z","consec"],1):
            ws.cell(row=r,column=c,value=e[k]).border=bd
        for c in range(13,16): ws.cell(row=r,column=c,value="").border=bd
    for c,w in enumerate([6,6,6,8,16,7,8,8,8,8,8,7,7,7,7],1):
        ws.column_dimensions[ws.cell(row=1,column=c).column_letter].width=w
    wb.save(fn); print("\n  Exported to: %s"%fn)

if __name__=="__main__":
    import time,json
    t0=time.time()
    sk,wd,pk,Z,M,log=cp_search()
    print("\n  Time: %.1fs"%(time.time()-t0))
    out=r"C:\Users\ming\Desktop\数模备赛\result_q3_cp_expected.xlsx"
    export_xlsx(sk,wd,pk,Z,M,log,out)
    sm=dict(problem="Task 3 (CP Expected-Value)",method="Constraint Programming + Expected Consumption + Safety Margins",
        path=" -> ".join(sk),Z=int(Z),M=int(M),
        work_days=dict(zip([n for n in sk if n in WORK_INFO],wd)),
        park_by_leg=pk,elapsed_s=round(time.time()-t0,1))
    with open(r"C:\Users\ming\Desktop\数模备赛\result_q3_cp_expected.json","w") as f: json.dump(sm,f,indent=2)
    print("  JSON: result_q3_cp_expected.json\nDone.")
