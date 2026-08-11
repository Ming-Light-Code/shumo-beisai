# -*- coding: utf-8 -*-
"""Q5: 多机多弹多目标协同干扰 —— 倒推候选生成 + 动态分支定界.

算法流程:
  第一步  单弹倒推候选 → 固定航向/速度的三弹套餐构造
  第二步  动态分支定界全局组合搜索
  第三步  细粒度最终复核

依赖: numpy, openpyxl, common.
"""
from __future__ import annotations

import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from dataclasses import dataclass
from math import atan2, cos, pi, sin, sqrt
from pathlib import Path
import random
import numpy as np

# ---- 从 common 导入所有共享常数 ----
from common import (
    GRAVITY, R_SMOKE, V_MISSILE, V_SINK, T_SMOKE_MAX,
    V_MIN, V_MAX, TARGET_CENTER, TARGET_POINT, ORIGIN,
    M0, FY_INIT, R_TARGET, H_TARGET,
)


# ============================================================
#  数据结构
# ============================================================
@dataclass
class Bomb:
    uav_id: int           # 无人机编号 (0-4)
    missile_id: int       # 目标导弹编号 (0-2)
    theta: float          # 航向角 (rad)
    v: float              # 飞行速度 (m/s)
    t_drop: float         # 投放时刻 (s)
    tau: float            # 起爆延迟 (s)


@dataclass
class Plan:
    bombs: list[Bomb]
    masks: np.ndarray     # (3, n_t) 布尔掩码
    value: float          # 累计遮蔽时长
    uav_id: int | None = None


# ============================================================
#  参数 (物理常数从 common 导入, 此处仅设 Q5 专属参数)
# ============================================================
def parameters() -> dict:
    """返回题目物理参数与初始条件."""
    p = {
        "g": GRAVITY,
        "vMissile": V_MISSILE,
        "Rsmoke": R_SMOKE,
        "smokeValidTime": T_SMOKE_MAX,
        "smokeSinkSpeed": V_SINK,
        "O": ORIGIN.copy(),
        "M0": M0.copy(),
        "FY": FY_INIT.copy(),
        "targetPoint": TARGET_POINT.copy(),
        "targetCenter": TARGET_CENTER.copy(),
        "targetR": R_TARGET,
        "targetH": H_TARGET,
    }
    p["tHit"] = np.linalg.norm(p["M0"] - p["O"], axis=1) / p["vMissile"]
    p["tEnd"] = float(np.max(p["tHit"]))
    return p


# ============================================================
#  工具函数
# ============================================================
def wrap_to_pi(a: float) -> float:
    return (a + pi) % (2 * pi) - pi


def mean_angle(values: list[float]) -> float:
    angle = atan2(float(np.mean(np.sin(values))), float(np.mean(np.cos(values))))
    return angle + 2 * pi if angle < 0 else angle


def missile_pos(t: float, missile_id: int, p: dict) -> np.ndarray:
    m0 = p["M0"][missile_id]
    return m0 + p["vMissile"] * t * (p["O"] - m0) / np.linalg.norm(p["O"] - m0)


def bomb_points(b: Bomb, p: dict) -> tuple[np.ndarray, np.ndarray, float]:
    f0 = p["FY"][b.uav_id]
    direction = np.array([cos(b.theta), sin(b.theta), 0.0])
    t_boom = b.t_drop + b.tau
    p_drop = f0 + b.v * b.t_drop * direction
    p_boom = p_drop + b.v * b.tau * direction + np.array([0.0, 0.0, -0.5 * p["g"] * b.tau ** 2])
    return p_drop, p_boom, t_boom


def sample_cylinder_two_rims(center: np.ndarray, radius: float, height: float,
                              n_theta: int) -> np.ndarray:
    """对圆柱上下底面圆周采样."""
    theta = np.arange(n_theta) * 2 * pi / n_theta
    bottom = np.column_stack((
        center[0] + radius * np.cos(theta),
        center[1] + radius * np.sin(theta),
        np.full(n_theta, center[2]),
    ))
    top = bottom.copy()
    top[:, 2] += height
    return np.vstack((bottom, top))


def segment_intersect_sphere(a: np.ndarray, b: np.ndarray, c: np.ndarray,
                              radius: float) -> bool:
    """线段-球体相交判定."""
    ab = b - a
    lam = float(np.dot(c - a, ab) / np.dot(ab, ab))
    lam = min(1.0, max(0.0, lam))
    q = a + lam * ab
    return bool(np.sum((q - c) ** 2) <= radius ** 2)


def target_points_blocked(m: np.ndarray, c: np.ndarray, radius: float,
                           pts: np.ndarray) -> np.ndarray:
    """判断每个采样点是否被烟幕球遮挡 (向量化)."""
    ab = pts - m
    ac = c - m
    lam = ab @ ac / np.sum(ab * ab, axis=1)
    lam = np.clip(lam, 0.0, 1.0)
    q = m + lam[:, None] * ab
    return np.sum((q - c) ** 2, axis=1) <= radius ** 2


# ============================================================
#  单弹遮蔽掩码
# ============================================================
def bomb_mask_center(theta: float, v: float, t_drop: float, tau: float,
                     f0: np.ndarray, missile_id: int, p: dict,
                     t_grid: np.ndarray) -> np.ndarray:
    """单弹中心线遮蔽掩码 (粗筛)."""
    b = Bomb(0, missile_id, theta, v, t_drop, tau)
    _, p_boom, t_boom = bomb_points_with_f0(b, f0, p)
    mask = np.zeros(t_grid.size, dtype=bool)
    if t_boom > p["tHit"][missile_id] or p_boom[2] < 0:
        return mask
    valid = np.flatnonzero(
        (t_grid >= t_boom) &
        (t_grid <= min(t_boom + p["smokeValidTime"], p["tHit"][missile_id]))
    )
    for k in valid:
        m = missile_pos(float(t_grid[k]), missile_id, p)
        c = p_boom + np.array([0.0, 0.0, -p["smokeSinkSpeed"] * (t_grid[k] - t_boom)])
        mask[k] = segment_intersect_sphere(m, p["targetPoint"], c, p["Rsmoke"])
    return mask


def bomb_points_with_f0(b: Bomb, f0: np.ndarray, p: dict
                        ) -> tuple[np.ndarray, np.ndarray, float]:
    direction = np.array([cos(b.theta), sin(b.theta), 0.0])
    t_boom = b.t_drop + b.tau
    p_drop = f0 + b.v * b.t_drop * direction
    p_boom = p_drop + b.v * b.tau * direction + np.array([0.0, 0.0, -0.5 * p["g"] * b.tau ** 2])
    return p_drop, p_boom, t_boom


def bomb_mask_rims(b: Bomb, p: dict, rim_pts: np.ndarray,
                   t_grid: np.ndarray) -> np.ndarray:
    """单弹圆柱边界遮蔽掩码 (精校)."""
    f0 = p["FY"][b.uav_id]
    _, p_boom, t_boom = bomb_points_with_f0(b, f0, p)
    mask = np.zeros(t_grid.size, dtype=bool)
    if t_boom > p["tHit"][b.missile_id] or p_boom[2] < 0:
        return mask
    valid = np.flatnonzero(
        (t_grid >= t_boom) &
        (t_grid <= min(t_boom + p["smokeValidTime"], p["tHit"][b.missile_id]))
    )
    for k in valid:
        m = missile_pos(float(t_grid[k]), b.missile_id, p)
        c = p_boom + np.array([0.0, 0.0, -p["smokeSinkSpeed"] * (t_grid[k] - t_boom)])
        mask[k] = bool(np.all(target_points_blocked(m, c, p["Rsmoke"], rim_pts)))
    return mask


# ============================================================
#  单弹候选倒推生成
# ============================================================
def inverse_decode(y: np.ndarray, f0: np.ndarray, missile_id: int,
                   p: dict) -> tuple[bool, tuple[float, float, float, float]]:
    """从 (t_boom, delta, s) 反解无人机飞行参数."""
    t_boom, delta, s = y
    if t_boom <= 0 or not (0 <= delta <= 20) or not (0 <= s <= 1):
        return False, (0, 0, 0, 0)
    t_center = t_boom + delta
    if t_center > p["tHit"][missile_id]:
        return False, (0, 0, 0, 0)
    mt = missile_pos(float(t_center), missile_id, p)
    line = mt + s * (p["targetPoint"] - mt)
    c_boom = line + np.array([0.0, 0.0, p["smokeSinkSpeed"] * delta])
    if c_boom[2] > f0[2] or c_boom[2] < 0:
        return False, (0, 0, 0, 0)
    tau = sqrt(2 * (f0[2] - c_boom[2]) / p["g"])
    t_drop = t_boom - tau
    if t_drop < 0 or tau < 0:
        return False, (0, 0, 0, 0)
    h = c_boom[:2] - f0[:2]
    v = float(np.linalg.norm(h) / t_boom)
    if not V_MIN <= v <= V_MAX:
        return False, (0, 0, 0, 0)
    theta = atan2(float(h[1]), float(h[0]))
    theta = theta + 2 * pi if theta < 0 else theta
    return True, (theta, v, t_drop, tau)


def generate_single_bomb_seeds(uav_id: int, missile_id: int, p: dict,
                               t_grid: np.ndarray, dt: float, n_seed: int,
                               k_keep: int, rng: random.Random) -> list[dict]:
    """为 (无人机, 导弹) 组合生成单弹候选种子."""
    seeds: list[dict] = []
    f0 = p["FY"][uav_id]
    t_hit = p["tHit"][missile_id]

    # ---- 阶段1: 倒推采样 ----
    for _ in range(n_seed):
        y = np.array([rng.random() * t_hit, rng.random() * 20, rng.random()])
        if y[0] + y[1] > t_hit:
            continue
        ok, x = inverse_decode(y, f0, missile_id, p)
        if not ok:
            continue
        theta, v, td, tau = x
        mask = bomb_mask_center(theta, v, td, tau, f0, missile_id, p, t_grid)
        value = float(np.sum(mask) * dt)
        if value > 0:
            seeds.append({"uav_id": uav_id, "missile_id": missile_id,
                          "theta": theta, "v": v, "t_drop": td, "tau": tau,
                          "mask": mask, "value": value})

    # ---- 阶段2: 直接随机采样 (倒推失败时) ----
    if not seeds:
        print(f"  FY{uav_id + 1}-M{missile_id + 1} 倒推采样无候选, 切换直接随机采样...")
        tau_max = sqrt(2 * f0[2] / p["g"])
        for _ in range(3 * n_seed):
            theta = 2 * pi * rng.random()
            v = V_MIN + (V_MAX - V_MIN) * rng.random()
            tau = tau_max * rng.random()
            td_max = max(0.0, t_hit - tau)
            if td_max <= 0:
                continue
            td = td_max * rng.random()
            mask = bomb_mask_center(theta, v, td, tau, f0, missile_id, p, t_grid)
            value = float(np.sum(mask) * dt)
            if value > 0:
                seeds.append({"uav_id": uav_id, "missile_id": missile_id,
                              "theta": theta, "v": v, "t_drop": td, "tau": tau,
                              "mask": mask, "value": value})

    # ---- 阶段3: 占位候选 ----
    if not seeds:
        print(f"  FY{uav_id + 1}-M{missile_id + 1} 无正向遮蔽候选, 使用可行占位候选...")
        theta = atan2(float(-f0[1]), float(-f0[0]))
        theta = theta + 2 * pi if theta < 0 else theta
        tau = min(3.0, sqrt(2 * f0[2] / p["g"]))
        td = max(0.0, min(1.0, t_hit - tau))
        mask = bomb_mask_center(theta, 100.0, td, tau, f0, missile_id, p, t_grid)
        seeds.append({"uav_id": uav_id, "missile_id": missile_id,
                      "theta": theta, "v": 100.0, "t_drop": td, "tau": tau,
                      "mask": mask, "value": float(np.sum(mask) * dt)})

    return sorted(seeds, key=lambda s: s["value"], reverse=True)[:k_keep]


# ============================================================
#  三弹套餐构造
# ============================================================
def make_bomb_from_seed(seed: dict, theta: float, v: float) -> Bomb:
    return Bomb(seed["uav_id"], seed["missile_id"], theta, v,
                seed["t_drop"], seed["tau"])


def evaluate_bomb_list(bombs: list[Bomb], p: dict, rim_pts: np.ndarray,
                       t_grid: np.ndarray, dt: float
                       ) -> tuple[np.ndarray, float]:
    """评估一组烟幕弹的联合遮蔽效果."""
    masks = np.zeros((3, t_grid.size), dtype=bool)
    for b in bombs:
        masks[b.missile_id] |= bomb_mask_rims(b, p, rim_pts, t_grid)
    return masks, float(np.sum(masks) * dt)


def build_uav_three_bomb_plans(uav_id: int, p: dict, rim_pts: np.ndarray,
                               t_grid: np.ndarray, dt: float, n_seed: int,
                               k_single: int, k_uav: int,
                               rng: random.Random) -> list[Plan]:
    """为单架无人机构造固定航向/速度的三弹套餐候选."""
    singles = [generate_single_bomb_seeds(uav_id, m, p, t_grid, dt,
                                          n_seed, k_single, rng)
               for m in range(3)]

    raw: list[Plan] = []
    for p1 in singles[0]:
        for p2 in singles[1]:
            for p3 in singles[2]:
                if (abs(wrap_to_pi(p1["theta"] - p2["theta"])) > 0.25 or
                    abs(wrap_to_pi(p1["theta"] - p3["theta"])) > 0.25):
                    continue
                if abs(p1["v"] - p2["v"]) > 12 or abs(p1["v"] - p3["v"]) > 12:
                    continue
                theta = mean_angle([p1["theta"], p2["theta"], p3["theta"]])
                v = float(np.mean([p1["v"], p2["v"], p3["v"]]))
                bombs = [make_bomb_from_seed(s, theta, v) for s in (p1, p2, p3)]
                bombs.sort(key=lambda b: b.t_drop)
                if any(bombs[j + 1].t_drop - bombs[j].t_drop < 1 for j in range(2)):
                    continue
                masks, value = evaluate_bomb_list(bombs, p, rim_pts, t_grid, dt)
                if value > 0:
                    raw.append(Plan(bombs, masks, value, uav_id))

    if not raw:
        bombs = [make_bomb_from_seed(singles[m][0], singles[m][0]["theta"],
                                     singles[m][0]["v"]) for m in range(3)]
        bombs.sort(key=lambda b: b.t_drop)
        for j in range(1, 3):
            if bombs[j].t_drop - bombs[j - 1].t_drop < 1:
                b = bombs[j]
                bombs[j] = Bomb(b.uav_id, b.missile_id, b.theta, b.v,
                                bombs[j - 1].t_drop + 1, b.tau)
        masks, value = evaluate_bomb_list(bombs, p, rim_pts, t_grid, dt)
        raw.append(Plan(bombs, masks, value, uav_id))

    return sorted(raw, key=lambda plan: plan.value, reverse=True)[:k_uav]


# ============================================================
#  动态分支定界
# ============================================================
def branch_bound_uav_plans(uav_sets: list[list[Plan]],
                           dt: float) -> tuple[Plan, dict]:
    """分支定界搜索五架无人机的最优套餐组合."""
    empty = np.zeros_like(uav_sets[0][0].masks, dtype=bool)
    best = Plan([], empty.copy(), -np.inf, None)
    info = {"visited": 0, "pruned": 0}

    def dfs(current: Plan, used: list[bool]) -> None:
        nonlocal best
        info["visited"] += 1
        if all(used):
            if current.value > best.value:
                best = Plan(current.bombs.copy(), current.masks.copy(),
                           current.value, None)
            return

        remaining = [u for u in range(5) if not used[u]]
        gains = []
        for u in remaining:
            gains.append(max(
                (float(np.sum(current.masks | cand.masks) -
                       np.sum(current.masks)) * dt
                 for cand in uav_sets[u]), default=0.0))
        upper = current.value + sum(gains)
        if upper <= best.value + 1e-12:
            info["pruned"] += 1
            return

        next_u = remaining[int(np.argmax(gains))]
        ranked = sorted(uav_sets[next_u],
                       key=lambda cand: float(
                           np.sum(current.masks | cand.masks) -
                           np.sum(current.masks)) * dt, reverse=True)
        for cand in ranked:
            new_masks = current.masks | cand.masks
            new_value = float(np.sum(new_masks) * dt)
            new_used = used.copy()
            new_used[next_u] = True
            ub2 = new_value
            for u in range(5):
                if not new_used[u]:
                    ub2 += max(float(np.sum(new_masks | c.masks) -
                                     np.sum(new_masks)) * dt
                              for c in uav_sets[u])
            if ub2 <= best.value + 1e-12:
                info["pruned"] += 1
                continue
            dfs(Plan(current.bombs + cand.bombs, new_masks, new_value, None),
                new_used)

    dfs(Plan([], empty, 0.0, None), [False] * 5)
    return best, info


# ============================================================
#  结果输出
# ============================================================
def mask_to_intervals(t_grid: np.ndarray, mask: np.ndarray,
                      dt: float) -> np.ndarray:
    """从布尔掩码提取连续遮蔽区间."""
    padded = np.r_[False, mask, False]
    start = np.flatnonzero(np.diff(padded.astype(int)) == 1)
    end = np.flatnonzero(np.diff(padded.astype(int)) == -1) - 1
    return np.array([[t_grid[a], t_grid[b], t_grid[b] - t_grid[a] + dt]
                     for a, b in zip(start, end)], dtype=float)


MISSILE_NAMES = ["M1", "M2", "M3"]
UAV_NAMES_Q5 = ["FY1", "FY2", "FY3", "FY4", "FY5"]


def write_result(best: Plan, p: dict, final_total: float,
                 final_times: np.ndarray, intervals: list[np.ndarray],
                 bomb_durations: list[float]) -> None:
    """输出 Excel 结果文件 (参照 A题 result3 格式)."""
    from openpyxl import Workbook

    wb = Workbook()

    # ================================================================
    #  Sheet1: 主结果表 (参照 A题 result3 格式)
    # ================================================================
    ws1 = wb.active
    ws1.title = "Sheet1"

    headers_main = ["无人机编号", "无人机运动方向", "无人机运动速度 (m/s)",
                    "烟幕弹编号", "烟幕弹投放点x坐标 (m)", "烟幕弹投放点y坐标 (m)",
                    "烟幕弹投放点z坐标 (m)", "烟幕弹起爆点x坐标 (m)",
                    "烟幕弹起爆点y坐标 (m)", "烟幕弹起爆点z坐标 (m)",
                    "有效干扰时间 (s)", "干扰的导弹"]
    ws1.append(headers_main)

    # 按 UAV 分组, 每组内按投放时刻排序
    bombs_sorted = sorted(best.bombs, key=lambda b: (b.uav_id, b.t_drop))
    uav_groups: dict[int, list[tuple[int, Bomb]]] = {}
    for idx, b in enumerate(bombs_sorted):
        uav_groups.setdefault(b.uav_id, []).append((idx, b))

    # 使用预计算的各弹单独遮蔽时长
    for uav_id in range(5):
        group = uav_groups.get(uav_id, [])
        for j, (orig_idx, b) in enumerate(group):
            pdp, pbm, tb = bomb_points(b, p)
            dir_deg = round(float(np.degrees(b.theta) % 360), 6)
            missile_name = MISSILE_NAMES[b.missile_id]
            dur_j = round(bomb_durations[orig_idx], 4)

            # 仅每组第一行填方向/速度
            dir_val = dir_deg if j == 0 else ""
            spd_val = round(float(b.v), 6) if j == 0 else ""

            ws1.append([UAV_NAMES_Q5[uav_id], dir_val, spd_val,
                       j + 1,
                       round(float(pdp[0]), 2), round(float(pdp[1]), 2), round(float(pdp[2]), 2),
                       round(float(pbm[0]), 2), round(float(pbm[1]), 2), round(float(pbm[2]), 2),
                       dur_j, missile_name])

    # 汇总行
    ws1.append([round(final_total, 4), "", "", "", "", "", "", "", "", "", "", ""])
    # 注释行
    ws1.append(["", "注：x轴为正北方向，逆时针方向为正，取值0~360度；", "", "", "", "", "", "", "", "", "", ""])

    # ================================================================
    #  Sheet2: 详细策略 (含全部运动参数)
    # ================================================================
    ws2 = wb.create_sheet("strategy_detail")
    headers_detail = ["烟幕弹序号", "无人机编号", "导弹编号",
                      "航向角_rad", "航向角_deg", "飞行速度",
                      "投放时刻", "起爆延迟", "起爆时刻",
                      "投放_X", "投放_Y", "投放_Z",
                      "起爆_X", "起爆_Y", "起爆_Z"]
    ws2.append(headers_detail)
    rows = []
    for number, b in enumerate(best.bombs, 1):
        pdp, pbm, tb = bomb_points(b, p)
        rows.append([number, b.uav_id + 1, b.missile_id + 1,
                     b.theta, np.degrees(b.theta), b.v,
                     b.t_drop, b.tau, tb] + list(pdp) + list(pbm))
    rows.sort(key=lambda r: (r[1], r[6]))
    for row in rows:
        ws2.append(row)

    # ================================================================
    #  Sheet3: 汇总
    # ================================================================
    ws3 = wb.create_sheet("summary")
    ws3.append(["累计总时长_s", "M1_s", "M2_s", "M3_s"])
    ws3.append([final_total, final_times[0], final_times[1], final_times[2]])

    # ================================================================
    #  Sheets 4-6: 各导弹遮蔽区间
    # ================================================================
    for m in range(3):
        ws = wb.create_sheet(f"M{m + 1}_intervals")
        ws.append(["起始时刻", "结束时刻", "持续时长"])
        for row_data in intervals[m]:
            ws.append(list(row_data))

    out = Path(__file__).parent / "result3.xlsx"
    wb.save(str(out))
    print(f"\n结果已保存至 {out}")


# ============================================================
#  主程序
# ============================================================
def main() -> None:
    p = parameters()
    rng = random.Random(1)

    dt_opt, dt_final = 0.05, 0.01
    k_single, k_uav, n_seed = 30, 20, 25000

    rim_opt = sample_cylinder_two_rims(p["targetCenter"], p["targetR"],
                                        p["targetH"], 120)
    rim_final = sample_cylinder_two_rims(p["targetCenter"], p["targetR"],
                                          p["targetH"], 360)

    t_grid = np.arange(0.0, p["tEnd"] + 0.5 * dt_opt, dt_opt)

    print("=" * 60)
    print("  问题五: 多机多弹多目标协同干扰优化")
    print("=" * 60)

    # ========= 第一步: 各无人机三弹套餐构造 =========
    print("\n第一步: 为每架无人机构造三弹套餐候选...")
    uav_sets = []
    for u in range(5):
        print(f"\n  构造 FY{u + 1} 候选...")
        plans = build_uav_three_bomb_plans(u, p, rim_opt, t_grid, dt_opt,
                                           n_seed, k_single, k_uav, rng)
        uav_sets.append(plans)
        print(f"  FY{u + 1}: 保留 {len(plans)} 个套餐, "
              f"最优评分 {plans[0].value:.6f} s")

    # ========= 第二步: 分支定界全局搜索 =========
    print("\n第二步: 动态分支定界全局组合搜索...")
    best, info = branch_bound_uav_plans(uav_sets, dt_opt)
    print(f"\n分支定界完成.")
    print(f"  搜索节点数: {info['visited']}")
    print(f"  剪枝节点数: {info['pruned']}")
    print(f"  搜索阶段最优评分: {best.value:.6f} s")

    # ========= 第三步: 细粒度最终复核 =========
    print("\n第三步: 细粒度最终复核...")
    t_final = np.arange(0.0, p["tEnd"] + 0.5 * dt_final, dt_final)
    final_masks = np.zeros((3, t_final.size), dtype=bool)
    for b in best.bombs:
        final_masks[b.missile_id] |= bomb_mask_rims(b, p, rim_final, t_final)

    final_times = np.sum(final_masks, axis=1) * dt_final
    final_total = float(np.sum(final_times))

    print("\n" + "=" * 60)
    print("  问题五最终结果")
    print("=" * 60)
    print(f"  M1 累计有效遮蔽时长: {final_times[0]:.6f} s")
    print(f"  M2 累计有效遮蔽时长: {final_times[1]:.6f} s")
    print(f"  M3 累计有效遮蔽时长: {final_times[2]:.6f} s")
    print(f"  全局累计遮蔽总时长: {final_total:.6f} s")

    intervals = []
    for m in range(3):
        iv = mask_to_intervals(t_final, final_masks[m], dt_final)
        intervals.append(iv)
        print(f"\n  M{m + 1} 有效遮蔽区间:")
        for row in iv:
            print(f"    [{row[0]:.2f}, {row[1]:.2f}]  持续 {row[2]:.4f} s")

    # 预计算各弹单独遮蔽时长 (复用第三步的细网格和采样点)
    print("\n  计算各弹单独遮蔽时长...")
    bomb_durations = []
    for b in best.bombs:
        dur_j = round(float(np.sum(bomb_mask_rims(b, p, rim_final, t_final))) * dt_final, 4)
        bomb_durations.append(dur_j)

    write_result(best, p, final_total, final_times, intervals, bomb_durations)


if __name__ == "__main__":
    main()
