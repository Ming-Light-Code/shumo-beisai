#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Q3 NLP: FY1 + 3 grenades vs M1.
Outer-loop NLP over (alpha, v) with inner fast greedy+refine from q3_solver."""

import numpy as np
import time, sys, os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import (
    G, VM, VS, RS, TMAX as TCLOUD, V_DRONE_MIN, V_DRONE_MAX, MIN_INTERVAL,
    MISSILES_INIT, DRONES_INIT, TARGET_SAMPLE_PTS,
    missile_position, drone_position, missile_flight_time,
    compute_detonation_pos, smoke_center_pos, check_occlusion_at_time,
)

M1_INIT = MISSILES_INIT["M1"]
FY1_INIT = DRONES_INIT["FY1"]
T_MAX = missile_flight_time(M1_INIT)
MAX_FALL = np.sqrt(2.0 * (FY1_INIT[2] + 10.0) / G)

# Fast optimization parameters (match q3_solver.py)
DT_OPT = 0.05
N_OPT = int(T_MAX / DT_OPT) + 1
T_OPT = np.linspace(0, T_MAX, N_OPT)
M1_OPT = np.array([missile_position(M1_INIT, t) for t in T_OPT])

# Fine verification
DT_FINE = 0.005
N_FINE = int(T_MAX / DT_FINE) + 1
T_FINE = np.linspace(0, T_MAX, N_FINE)
M1_FINE = np.array([missile_position(M1_INIT, t) for t in T_FINE])


def quick_occlusion(alpha, v, td, tb):
    """Fast single-grenade occlusion (same approach as q3_solver.py)."""
    if tb >= T_MAX: return 0.0
    t0_idx = int(tb / DT_OPT)
    t1_idx = min(N_OPT, int(min(tb + TCLOUD, T_MAX) / DT_OPT) + 1)
    if t0_idx >= t1_idx: return 0.0
    det_pos = compute_detonation_pos(FY1_INIT, alpha, v, td, tb)
    cnt = 0
    for i in range(t0_idx, t1_idx):
        sc = smoke_center_pos(det_pos, tb, T_OPT[i])
        if check_occlusion_at_time(M1_OPT[i], sc, TARGET_SAMPLE_PTS):
            cnt += 1
    return cnt * DT_OPT


def quick_union(alpha, v, triplets):
    """Fast union occlusion for triplets (same approach as q3_solver.py)."""
    if not triplets: return 0.0
    t_start = min(tb for _, tb in triplets)
    t_end = T_MAX
    i0 = int(t_start / DT_OPT)
    i1 = min(N_OPT, int(t_end / DT_OPT) + 1)
    if i0 >= i1: return 0.0
    n = i1 - i0
    union = np.zeros(n, dtype=bool)
    for td, tb in triplets:
        det_pos = compute_detonation_pos(FY1_INIT, alpha, v, td, tb)
        for idx in range(i0, i1):
            t = T_OPT[idx]
            if t < tb or t > tb + TCLOUD: continue
            sc = smoke_center_pos(det_pos, tb, t)
            if check_occlusion_at_time(M1_OPT[idx], sc, TARGET_SAMPLE_PTS):
                union[idx - i0] = True
    return float(np.sum(union)) * DT_OPT



# Fast occlusion with reduced target points for seed generation
TARGET_FAST = np.array([
    [0, 200, 5], [7, 200, 5], [-7, 200, 5],
    [0, 207, 5], [0, 193, 5], [0, 200, 10], [0, 200, 0],
], dtype=float)

def quick_occlusion_fast(alpha, v, td, tb):
    """Fast occlusion using only 7 key target points."""
    if tb >= T_MAX: return 0.0
    t0_idx = int(tb / DT_OPT)
    t1_idx = min(N_OPT, int(min(tb + TCLOUD, T_MAX) / DT_OPT) + 1)
    if t0_idx >= t1_idx: return 0.0
    det_pos = compute_detonation_pos(FY1_INIT, alpha, v, td, tb)
    cnt = 0
    for i in range(t0_idx, t1_idx):
        sc = smoke_center_pos(det_pos, tb, T_OPT[i])
        if check_occlusion_at_time(M1_OPT[i], sc, TARGET_FAST):
            cnt += 1
    return cnt * DT_OPT


# ======================================================================
#   Vectorized occlusion (massive speedup over per-timestep Python loop)
# ======================================================================

def quick_occlusion_vec(alpha, v, td, tb, target_pts=None):
    """Vectorized single-grenade occlusion check."""
    if target_pts is None: target_pts = TARGET_SAMPLE_PTS
    if tb >= T_MAX: return 0.0
    t0_idx = int(tb / DT_OPT)
    t1_idx = min(N_OPT, int(min(tb + TCLOUD, T_MAX) / DT_OPT) + 1)
    if t0_idx >= t1_idx: return 0.0
    det_pos = compute_detonation_pos(FY1_INIT, alpha, v, td, tb)

    n = t1_idx - t0_idx
    M_block = M1_OPT[t0_idx:t1_idx]                     # (n,3)
    zc = det_pos[2] - VS * (T_OPT[t0_idx:t1_idx] - tb)  # (n,)
    C = np.column_stack([np.full(n, det_pos[0]),
                         np.full(n, det_pos[1]), zc])    # (n,3)

    dc = C - M_block                                      # (n,3)
    dc_norm = np.linalg.norm(dc, axis=1)                  # (n,)
    valid = (dc_norm > RS) & (dc_norm > 1e-9)
    result = np.ones(n, dtype=bool)
    theta_s = np.where(valid, np.arcsin(RS / dc_norm), np.pi)

    for pt in target_pts:
        vp = pt - M_block                                 # (n,3)
        vp_norm = np.linalg.norm(vp, axis=1)              # (n,)
        v_ok = vp_norm > 1e-9
        # Only compute for valid timesteps
        ang = np.full(n, np.pi)
        idx = valid & v_ok
        if np.any(idx):
            vpu = vp[idx] / vp_norm[idx, np.newaxis]
            dcu = dc[idx] / dc_norm[idx, np.newaxis]
            cos_a = np.clip(np.sum(vpu * dcu, axis=1), -1.0, 1.0)
            ang[idx] = np.arccos(cos_a)
        result &= (ang <= theta_s)
    return np.sum(result) * DT_OPT


def quick_union_vec(alpha, v, triplets):
    """Vectorized union occlusion for triplets."""
    if not triplets: return 0.0
    t_start = min(tb for _, tb in triplets)
    t_end = T_MAX
    i0 = int(t_start / DT_OPT)
    i1 = min(N_OPT, int(t_end / DT_OPT) + 1)
    if i0 >= i1: return 0.0
    n = i1 - i0
    union = np.zeros(n, dtype=bool)
    M_block = M1_OPT[i0:i1]                               # (n,3)
    ts = T_OPT[i0:i1]                                     # (n,)

    for td, tb in triplets:
        if tb >= T_MAX: continue
        det_pos = compute_detonation_pos(FY1_INIT, alpha, v, td, tb)
        zc = det_pos[2] - VS * np.maximum(0, ts - tb)     # (n,)
        C = np.column_stack([np.full(n, det_pos[0]),
                             np.full(n, det_pos[1]), zc]) # (n,3)

        dc = C - M_block                                   # (n,3)
        dc_norm = np.linalg.norm(dc, axis=1)               # (n,)
        valid = (dc_norm > RS) & (dc_norm > 1e-9)
        theta_s = np.where(valid, np.arcsin(RS / dc_norm), np.pi)
        active = (ts >= tb) & (ts <= tb + TCLOUD)

        occ = np.zeros(n, dtype=bool)
        for pt in TARGET_SAMPLE_PTS:
            vp = pt - M_block                              # (n,3)
            vp_norm = np.linalg.norm(vp, axis=1)
            v_ok = vp_norm > 1e-9
            idx = valid & v_ok & active
            ang = np.full(n, np.pi)
            if np.any(idx):
                vpu = vp[idx] / vp_norm[idx, np.newaxis]
                dcu = dc[idx] / dc_norm[idx, np.newaxis]
                cos_a = np.clip(np.sum(vpu * dcu, axis=1), -1.0, 1.0)
                ang[idx] = np.arccos(cos_a)
            occ |= (ang <= theta_s) & active
        union |= occ
    return float(np.sum(union)) * DT_OPT



# Fast union with 7 target points - for optimization phases only
def quick_union_fast(alpha, v, triplets):
    if not triplets: return 0.0
    t_start = min(tb for _, tb in triplets)
    t_end = T_MAX
    i0 = int(t_start / DT_OPT)
    i1 = min(N_OPT, int(t_end / DT_OPT) + 1)
    if i0 >= i1: return 0.0
    n = i1 - i0
    union = np.zeros(n, dtype=bool)
    M_block = M1_OPT[i0:i1]
    ts = T_OPT[i0:i1]

    for td, tb in triplets:
        if tb >= T_MAX: continue
        det_pos = compute_detonation_pos(FY1_INIT, alpha, v, td, tb)
        zc = det_pos[2] - VS * np.maximum(0, ts - tb)
        C = np.column_stack([np.full(n, det_pos[0]),
                             np.full(n, det_pos[1]), zc])
        dc = C - M_block
        dc_norm = np.linalg.norm(dc, axis=1)
        valid = (dc_norm > RS) & (dc_norm > 1e-9)
        theta_s = np.where(valid, np.arcsin(np.clip(RS / dc_norm, -1, 1)), np.pi)
        active = (ts >= tb) & (ts <= tb + TCLOUD)
        occ = np.zeros(n, dtype=bool)
        for pt in TARGET_FAST:
            vp = pt - M_block
            vp_norm = np.linalg.norm(vp, axis=1)
            v_ok = vp_norm > 1e-9
            idx = valid & v_ok & active
            ang = np.full(n, np.pi)
            if np.any(idx):
                vpu = vp[idx] / vp_norm[idx, np.newaxis]
                dcu = dc[idx] / dc_norm[idx, np.newaxis]
                cos_a = np.clip(np.sum(vpu * dcu, axis=1), -1.0, 1.0)
                ang[idx] = np.arccos(cos_a)
            occ |= (ang <= theta_s) & active
        union |= occ
    return float(np.sum(union)) * DT_OPT


# Fast single-grenade occlusion for use in coordinate descent
def quick_occlusion_fast_vec(alpha, v, td, tb):
    return quick_occlusion_vec(alpha, v, td, tb, TARGET_FAST)

def generate_candidates(alpha, v, n=20):
    """Generate single-grenade candidates for given (alpha, v)."""
    cands = []
    for td in np.linspace(0, 15, 15):
        for tb in np.linspace(td + 0.1, min(td + 15, T_MAX - 1), 15):
            oc = quick_occlusion_vec(alpha, v, td, tb, TARGET_FAST)
            if oc > 0.3:
                det_pos = compute_detonation_pos(FY1_INIT, alpha, v, td, tb)
                cands.append((td, tb, oc, det_pos))
    cands.sort(key=lambda c: -c[2])
    filtered = []
    for c in cands:
        td, tb, oc, _ = c
        dup = False
        for f in filtered:
            if abs(td - f[0]) < 0.5 and abs(tb - f[1]) < 0.8:
                dup = True; break
        if not dup: filtered.append(c)
        if len(filtered) >= n: break
    return filtered


def greedy_fill(alpha, v, candidates, max_n=3):
    """Greedily add grenades to maximize union occlusion."""
    selected = []
    used = set()
    for _ in range(max_n):
        best_occ = quick_union_fast(alpha, v, selected)
        best_idx = None
        for i, (td, tb, oc, _) in enumerate(candidates):
            if i in used: continue
            ok = True
            for prev_td, prev_tb in selected:
                if td < prev_td + MIN_INTERVAL:
                    ok = False; break
            if not ok: continue
            trial = selected + [(td, tb)]
            trial_occ = quick_union_fast(alpha, v, trial)
            if trial_occ > best_occ + 0.001:
                best_occ = trial_occ
                best_idx = i
        if best_idx is not None:
            selected.append((candidates[best_idx][0], candidates[best_idx][1]))
            used.add(best_idx)
        else:
            break
    return selected, quick_union_vec(alpha, v, selected)


def local_refine(alpha, v, triplets):
    """Coordinate-descent refinement (NLP) for given (alpha, v)."""
    best = [list(t) for t in triplets]
    best_occ = quick_union_fast(alpha, v, best)
    for step in [0.1, 0.05, 0.01, 0.005]:
        improved = True
        while improved:
            improved = False
            for i in range(len(best)):
                for j in [0, 1]:
                    for sgn in [-1, 1]:
                        new = [list(t) for t in best]
                        new[i][j] += sgn * step
                        if j == 0:
                            if new[i][0] < 0: continue
                            if i > 0 and new[i][0] < new[i-1][0] + MIN_INTERVAL: continue
                        if j == 1:
                            if new[i][1] <= new[i][0]: continue
                            if new[i][1] >= T_MAX: continue
                            zb = FY1_INIT[2] - 0.5*G*(new[i][1]-new[i][0])**2
                            if zb < -10: continue
                        occ = quick_union_fast(alpha, v, new)
                        if occ > best_occ + 1e-8:
                            best = new; best_occ = occ; improved = True
            if not improved: break
    return best, best_occ
# ======================================================================
#   NLP: Optimize over all 8 variables (alpha, v, td1, tb1, td2, tb2, td3, tb3)
# ======================================================================

def build_x(alpha, v, triplets):
    """Convert (alpha, v, triplets) to 8-vector x."""
    x = np.zeros(8)
    x[0], x[1] = alpha, v
    for k in range(min(3, len(triplets))):
        x[2+2*k], x[3+2*k] = triplets[k][0], triplets[k][1]
    return x


def full_coordinate_descent(x0):
    """Coordinate descent on all 8 variables (the NLP core)."""
    xb = np.array(x0, dtype=float)
    fb = eval_8var(xb)
    for step in [0.5, 0.2, 0.1, 0.05, 0.02, 0.01, 0.005]:
        improved = True
        while improved:
            improved = False
            for d in range(8):
                for sgn in [-1, 1]:
                    xt = xb.copy()
                    xt[d] += sgn * step
                    xt[0] %= (2.0 * np.pi)
                    xt[1] = np.clip(xt[1], V_DRONE_MIN, V_DRONE_MAX)
                    xt[2] = max(0.0, xt[2])
                    xt[3] = max(xt[2] + 0.01, min(xt[3], T_MAX - 0.01))
                    xt[4] = max(xt[2] + MIN_INTERVAL, xt[4])
                    xt[5] = max(xt[4] + 0.01, min(xt[5], T_MAX - 0.01))
                    xt[6] = max(xt[4] + MIN_INTERVAL, xt[6])
                    xt[7] = max(xt[6] + 0.01, min(xt[7], T_MAX - 0.01))
                    for k in range(3):
                        dti = xt[3+2*k] - xt[2+2*k]
                        max_dt = MAX_FALL
                        if dti > max_dt: xt[3+2*k] = xt[2+2*k] + max_dt
                    ft = eval_8var(xt)
                    if ft > fb + 1e-8:
                        xb = xt; fb = ft; improved = True
    return xb, fb


def eval_8var(x):
    """Evaluate union occlusion for 8-vector x."""
    alpha, v = x[0], x[1]
    t = [(x[2], x[3]), (x[4], x[5]), (x[6], x[7])]
    return quick_union_fast(alpha, v, t)




def eval_8var_full(x):
    """Evaluate union occlusion for 8-vector x (full 160 pts)."""
    alpha, v = x[0], x[1]
    t = [(x[2], x[3]), (x[4], x[5]), (x[6], x[7])]
    return quick_union_fast(alpha, v, t)

def verify_8var_fast(x):
    alpha, v = x[0], x[1]
    tds = [x[2], x[4], x[6]]
    tbs = [x[3], x[5], x[7]]
    if min(tbs) >= T_MAX: return 0.0
    i0 = max(0, int(min(tbs) / DT_OPT))
    i1 = min(N_OPT, int(T_MAX / DT_OPT) + 1)
    if i0 >= i1: return 0.0
    n = i1 - i0
    union = np.zeros(n, dtype=bool)
    for k in range(3):
        tb_k = tbs[k]
        if tb_k >= T_MAX: continue
        det_pos = compute_detonation_pos(FY1_INIT, alpha, v, tds[k], tb_k)
        for i in range(i0, i1):
            t = T_OPT[i]
            if t < tb_k or t > tb_k + TCLOUD: continue
            if not union[i - i0]:
                sc = smoke_center_pos(det_pos, tb_k, t)
                if check_occlusion_at_time(M1_OPT[i], sc, TARGET_SAMPLE_PTS):
                    union[i - i0] = True
    return float(np.sum(union)) * DT_OPT

def verify_8var(x):
    """Fine-DT verification of 8-vector x."""
    alpha, v = x[0], x[1]
    triplets = [(x[2], x[3]), (x[4], x[5]), (x[6], x[7])]

    # Fine union
    tds = [x[2], x[4], x[6]]
    tbs = [x[3], x[5], x[7]]
    t_begin = min(tbs)
    if t_begin >= T_MAX: return 0.0, [0, 0, 0]
    i0 = max(0, int(t_begin / DT_FINE))
    i1 = min(N_FINE, int(T_MAX / DT_FINE) + 1)
    if i0 >= i1: return 0.0, [0, 0, 0]
    n = i1 - i0
    union = np.zeros(n, dtype=bool)
    for k in range(3):
        tb_k = tbs[k]
        if tb_k >= T_MAX: continue
        det_pos = compute_detonation_pos(FY1_INIT, alpha, v, tds[k], tb_k)
        for i in range(i0, i1):
            t = T_FINE[i]
            if t < tb_k or t > tb_k + TCLOUD: continue
            if not union[i - i0]:
                sc = smoke_center_pos(det_pos, tb_k, t)
                if check_occlusion_at_time(M1_FINE[i], sc, TARGET_SAMPLE_PTS):
                    union[i - i0] = True
    occ_union = float(np.sum(union)) * DT_FINE

    # Individual occlusions
    ocs = []
    for k in range(3):
        tb_k = tbs[k]
        if tb_k >= T_MAX:
            ocs.append(0.0); continue
        det_pos = compute_detonation_pos(FY1_INIT, alpha, v, tds[k], tb_k)
        cnt = 0
        i0_k = max(0, int(tb_k / DT_FINE))
        i1_k = min(N_FINE, int(min(tb_k + TCLOUD, T_MAX) / DT_FINE) + 1)
        for i in range(i0_k, i1_k):
            sc = smoke_center_pos(det_pos, tb_k, T_FINE[i])
            if check_occlusion_at_time(M1_FINE[i], sc, TARGET_SAMPLE_PTS):
                cnt += 1
        ocs.append(cnt * DT_FINE)

    return occ_union, ocs
# ======================================================================
#   Main multi-start NLP
# ======================================================================



def full_coordinate_descent_final(x0):
    xb = np.array(x0, dtype=float)
    fb = eval_8var_full(xb)
    for step in [0.1, 0.05, 0.02, 0.01, 0.005]:
        improved = True
        while improved:
            improved = False
            for d in range(8):
                for sgn in [-1, 1]:
                    xt = xb.copy(); xt[d] += sgn * step
                    xt[0] %= (2.0 * np.pi)
                    xt[1] = np.clip(xt[1], V_DRONE_MIN, V_DRONE_MAX)
                    xt[2] = max(0.0, xt[2])
                    xt[3] = max(xt[2] + 0.01, min(xt[3], T_MAX - 0.01))
                    xt[4] = max(xt[2] + MIN_INTERVAL, xt[4])
                    xt[5] = max(xt[4] + 0.01, min(xt[5], T_MAX - 0.01))
                    xt[6] = max(xt[4] + MIN_INTERVAL, xt[6])
                    xt[7] = max(xt[6] + 0.01, min(xt[7], T_MAX - 0.01))
                    for k in range(3):
                        dti = xt[3+2*k] - xt[2+2*k]
                        if dti > MAX_FALL: xt[3+2*k] = xt[2+2*k] + MAX_FALL
                    ft = eval_8var_full(xt)
                    if ft > fb + 1e-8: xb = xt; fb = ft; improved = True
    return xb, fb

def solve_q3_nlp():
    print("=" * 70)
    print("  Q3 NLP: Outer-loop (a,v) + Inner greedy/refine")
    print(f"  DT_OPT={DT_OPT}s  DT_FINE={DT_FINE}s  T_MAX={T_MAX:.2f}s")
    print("=" * 70)
    t_start = time.time()

    # Search (alpha, v) grid
    print("\n  [1] Searching (alpha, v) grid...")
    results = []

    for a_deg in [172, 174, 176, 177, 178, 179, 180, 181, 182, 183, 184, 186, 188]:
        alpha = np.radians(a_deg)
        for v_val in [130, 135, 140]:
            # Generate candidates
            cands = generate_candidates(alpha, v_val, n=20)
            # Greedy fill
            triplets, occ_greedy = greedy_fill(alpha, v_val, cands, 3)
            # Local refine
            ref_triplets, occ_ref = local_refine(alpha, v_val, triplets)
            # Build 8-vector and run full coordinate descent
            x0 = build_x(alpha, v_val, ref_triplets)
            x_nlp, occ_nlp = full_coordinate_descent(x0)

            # Final full-union polish of CD result
            x_polish, occ_polish = full_coordinate_descent_final(x_nlp)
            results.append({"alpha": np.degrees(x_polish[0]), "v": x_polish[1],
                          "x": x_polish, "occ_opt": occ_polish})

    # Sort by fast-union score, then pick best for full-union polish
    results.sort(key=lambda r: -r["occ_opt"])
    print(f"  Searched {len(results)} (alpha, v) pairs ({time.time()-t_start:.1f}s)")

    # Thorough full-union CD on top-3 only (for solution quality)
    print("  [1b] Full-union CD polish on top-3 results...")
    t_polish = time.time()
    for r in results[:3]:
        x0 = r["x"]
        # Full CD with thorough step sizes
        xb = np.array(x0, dtype=float)
        fb = quick_union_vec(xb[0], xb[1], [(xb[2],xb[3]),(xb[4],xb[5]),(xb[6],xb[7])])
        for step in [0.1, 0.05, 0.02, 0.01, 0.005, 0.002]:
            improved = True
            while improved:
                improved = False
                for d in range(8):
                    for sgn in [-1, 1]:
                        xt = xb.copy(); xt[d] += sgn * step
                        xt[0] %= (2.0 * np.pi)
                        xt[1] = np.clip(xt[1], V_DRONE_MIN, V_DRONE_MAX)
                        xt[2] = max(0.0, xt[2])
                        xt[3] = max(xt[2] + 0.01, min(xt[3], T_MAX - 0.01))
                        xt[4] = max(xt[2] + MIN_INTERVAL, xt[4])
                        xt[5] = max(xt[4] + 0.01, min(xt[5], T_MAX - 0.01))
                        xt[6] = max(xt[4] + MIN_INTERVAL, xt[6])
                        xt[7] = max(xt[6] + 0.01, min(xt[7], T_MAX - 0.01))
                        for k in range(3):
                            dti = xt[3+2*k] - xt[2+2*k]
                            if dti > MAX_FALL: xt[3+2*k] = xt[2+2*k] + MAX_FALL
                        ft = quick_union_vec(xt[0], xt[1], [(xt[2],xt[3]),(xt[4],xt[5]),(xt[6],xt[7])])
                        if ft > fb + 1e-8: xb = xt; fb = ft; improved = True
        r["x"] = xb
        r["occ_opt"] = fb
    print(f"  Polish done ({time.time()-t_polish:.1f}s)")
    results.sort(key=lambda r: -r["occ_opt"])

    print("\n  [2] Fine-DT verification of top results...")
    for r in results:
        occ_f, ocs = verify_8var(r["x"])
        r["occ_fine"] = occ_f
        r["ocs_fine"] = ocs

    results.sort(key=lambda r: -r["occ_fine"])

    print(f"\n  Top results (fine DT):")
    for i, r in enumerate(results[:10]):
        a = r["alpha"]; v = r["v"]
        occ = r["occ_fine"]
        o1, o2, o3 = r["ocs_fine"]
        print(f"    {i+1:2d}. a={a:7.2f}deg v={v:6.1f}"
              f" occ={occ:.4f}s (G1={o1:.3f} G2={o2:.3f} G3={o3:.3f})")

    best = results[0]
    print(f"\n  Best overall: {best['occ_fine']:.6f}s")
    print(f"  Total time: {time.time()-t_start:.1f}s")
    print("=" * 70)
    return best["x"], best["occ_fine"], results


# ======================================================================
#   Export and main
# ======================================================================

def export_result1(best_x):
    """Export to result1.xlsx."""
    import openpyxl
    alpha, v = best_x[0], best_x[1]
    td = [best_x[2], best_x[4], best_x[6]]
    tb = [best_x[3], best_x[5], best_x[7]]
    occ_union, ocs = verify_8var(best_x)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投放策略"
    hdrs = ["烟幕干扰弹编号","无人机编号","无人机飞行方向(deg)","无人机飞行速度(m/s)",
            "投放时刻(s)","起爆时刻(s)","起爆点X(m)","起爆点Y(m)","起爆点Z(m)","有效遮蔽时长(s)"]
    for j, h in enumerate(hdrs): ws.cell(row=1, column=j+1, value=h)
    for k in range(3):
        row = k + 2
        det_pos = compute_detonation_pos(FY1_INIT, alpha, v, td[k], tb[k])
        ws.cell(row=row, column=1, value=f"G{k+1}")
        ws.cell(row=row, column=2, value="FY1")
        ws.cell(row=row, column=3, value=round(np.degrees(alpha), 6))
        ws.cell(row=row, column=4, value=round(v, 6))
        ws.cell(row=row, column=5, value=round(td[k], 6))
        ws.cell(row=row, column=6, value=round(tb[k], 6))
        ws.cell(row=row, column=7, value=round(float(det_pos[0]), 3))
        ws.cell(row=row, column=8, value=round(float(det_pos[1]), 3))
        ws.cell(row=row, column=9, value=round(float(det_pos[2]), 3))
        ws.cell(row=row, column=10, value=round(ocs[k], 6))
    ws.cell(row=6, column=1, value="合计")
    ws.cell(row=6, column=10, value=round(occ_union, 6))

    ws_info = wb.create_sheet("信息")
    info = [("无人机编号","FY1"),("飞行方向(deg)",round(np.degrees(alpha),6)),
            ("飞行速度(m/s)",round(v,6)),("总有效遮蔽时长(s)",round(occ_union,6)),
            ("方法","非线性规划-坐标下降"),("G1遮蔽时长(s)",round(ocs[0],6)),
            ("G2遮蔽时长(s)",round(ocs[1],6)),("G3遮蔽时长(s)",round(ocs[2],6))]
    ws_info.cell(row=1, column=1, value="名称"); ws_info.cell(row=1, column=2, value="值")
    for i, (k, val) in enumerate(info):
        ws_info.cell(row=i+2, column=1, value=k); ws_info.cell(row=i+2, column=2, value=val)

    out_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "result1.xlsx")
    wb.save(out_path)
    print(f"\n  result1.xlsx saved to {out_path}")
    return out_path


if __name__ == "__main__":
    best_x, best_occ, all_results = solve_q3_nlp()

    alpha, v = best_x[0], best_x[1]
    td = [best_x[2], best_x[4], best_x[6]]
    tb = [best_x[3], best_x[5], best_x[7]]

    print("\n" + "=" * 70)
    print("  Q3 FINAL RESULT (NLP)")
    print("=" * 70)
    print(f"  FY1: heading = {np.degrees(alpha):.6f}deg, speed = {v:.4f} m/s")
    for k in range(3):
        det_pos = compute_detonation_pos(FY1_INIT, alpha, v, td[k], tb[k])
        print(f"  G{k+1}: td={td[k]:.6f}s  tb={tb[k]:.6f}s"
              f"  det=({det_pos[0]:.1f},{det_pos[1]:.1f},{det_pos[2]:.1f})")
    occ_f, ocs = verify_8var(best_x)
    print(f"  Union occlusion: {occ_f:.6f}s (G1={ocs[0]:.4f} G2={ocs[1]:.4f} G3={ocs[2]:.4f})")
    print("=" * 70)

    export_result1(best_x)